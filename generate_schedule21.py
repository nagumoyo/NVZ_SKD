#!/usr/bin/env python3
# === generate_schedule21.py ===

import pandas as pd
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ===


def clean_cell(text):
    """Remove invisible characters and trim whitespace."""
    return re.sub(
        r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", str(text)
    ).strip()


def read_and_clean_schedule(schedule_file):
    """Load CSV and drop rows that are empty, contain 'OB', or have emp_no 00099xxx."""
    df = pd.read_csv(schedule_file, header=None, dtype=str).fillna("")
    print(f"📥 元データ行数: {len(df)}")

    def is_useless_row(row):
        name = str(row[0]).strip()
        emp_no = str(row[15]).strip() if len(row) > 15 else ""
        if all(str(cell).strip() == "" for cell in row):
            return True
        if "OB" in name:
            return True
        if re.match(r"^00099\d{3}$", emp_no):
            return True
        return False

    cleaned_df = df[~df.apply(is_useless_row, axis=1)].reset_index(drop=True)
    print(f"🧹 フィルタ後行数: {len(cleaned_df)}")
    return cleaned_df


def find_header_rows(df):
    hdrs = []
    for i in range(len(df) - 1):
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    print(f"🔍 検出されたヘッダー行: {hdrs}")
    return hdrs


def slice_blocks(df):
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        d = h + 1
        vals = [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    d = r
                    break
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]
        blocks.append((h, d, end, date_cols))
    print(f"📦 ブロック数: {len(blocks)}")
    return blocks


# Excel styles
HIGHLIGHT = PatternFill(fill_type="solid", fgColor="FFEE99")
PH_HIGHLIGHT = PatternFill(fill_type="solid", fgColor="9393FF")
DEFAULT_GREY = PatternFill(fill_type="solid", fgColor="DDDDDD")
DOUBLE = Side(border_style="double", color="000000")


def write_to_excel(records, emp_aff_map, out_xlsx):
    wb = Workbook()
    ws = wb.active
    row_num = 1
    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}

    for rec in records:
        block_aff = rec["aff"]
        for j, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=j, value=val)
            wrap = not bool(re.fullmatch(r"0[0-9]{1,}-[0-9]+-[0-9]{4}", val))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=wrap
            )
            cell.border = Border(top=DOUBLE, bottom=DOUBLE)
            if j == 30 and val == "PH3":
                cell.fill = PH_HIGHLIGHT
        for j, val in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=row_num + 1, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = DEFAULT_GREY
        for j, val in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num + 2, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        for j, names in enumerate(rec["onb"], start=1):
            cell = ws.cell(row=row_num + 3, column=j, value="\n".join(names))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            for name in names:
                emp = name_to_emp.get(name)
                if emp and emp_aff_map.get(emp) == block_aff:
                    cell.fill = HIGHLIGHT
                    break
        row_num += 4
    wb.save(out_xlsx)
    print(f"📤 Excelファイル書き出し完了: {out_xlsx}")


# Main


def run(schedule_file, emp_file):
    df = read_and_clean_schedule(schedule_file).map(clean_cell)
    emp_df = pd.read_csv(emp_file, header=None, dtype=str).fillna("")
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()

    blocks = slice_blocks(df)
    if not blocks:
        print("⚠️ ブロックが検出されませんでした。処理を中止します。")
        return

    global_dates = blocks[0][3]
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        raw[0] = f"{surname}{two}" if matched else raw[0]
        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        col8 = emp_col8_map.get(code, "")
        m = re.search(r"(\d+)", col8)
        if m:
            hdr[29] = f"PH{m.group(1)}"
        hdr[30] = rec_aff
        hdr = [
            re.sub(
                r"電話番号",
                "電話",
                re.sub(
                    r"社員番号",
                    "職番",
                    re.sub(r"PE([0-9]{6})", r"\1", re.sub(r"PE有効期限", "PE", v)),
                ),
            )
            for v in hdr
        ]
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = [
            [
                clean_cell(df.iat[r2, j])
                for r2 in range(d + 1, end)
                if clean_cell(df.iat[r2, j])
            ]
            for j in dates
        ]
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))
        records.append(
            {
                "emp_no": code,
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )

    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))
    for i, rec in enumerate(records):
        onb = []
        for idx, entries in enumerate(rec["full_entries"]):
            flights = [e for e in entries if re.match(r"^[0-9]", e)]
            names = []
            for j, other in enumerate(records):
                if i != j and any(f in other["full_entries"][idx] for f in flights):
                    names.append(other["hdr"][0])
            uniq = []
            [uniq.append(n) for n in names if n not in uniq]
            onb.append(uniq)
        rec["onb"] = onb
    seen = set()
    uniq = []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = uniq
    records.sort(
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        )
    )
    with open("formatted_schedule.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec["onb"]])
    print("📄 CSVファイル書き出し完了: formatted_schedule.csv")
    write_to_excel(records, emp_aff_map, "formatted_schedule20.xlsx")
    print("✅ 全処理完了")
    return "formatted_schedule.csv", "formatted_schedule20.xlsx"


import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("generate_schedule.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--emp", default="emp_no.csv")
    a = p.parse_args()
    run(a.schedule, a.emp)
