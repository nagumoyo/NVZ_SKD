#!/usr/bin/env python3
# === generate_schedule35d.py ===

import pandas as pd
import re
import csv
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === フィルター用：M列（hdr[12]）から養成期の数値を拾う ==================
# 0-based index: A=0, ..., L=11, M=12
HDR_IDX_M = 12


def _phase_from_hdr_m(hdr):
    """
    ヘッダー配列 hdr の M 列（hdr[12]）を見て、
    2～3桁の純粋な数字なら int で返す。そうでなければ None。
    """
    if not isinstance(hdr, (list, tuple)) or len(hdr) <= HDR_IDX_M:
        return None
    s = str(hdr[HDR_IDX_M] or "").strip()
    return int(s) if s.isdigit() and 2 <= len(s) <= 3 else None


# ===================================================================


# === 前回Excel差分ハイライト (emp_no × day_index 比較) ==============
import re as _re_for_prevmap
from openpyxl import load_workbook as _load_wb_for_prevmap
from openpyxl.styles import PatternFill as _PatternFill_for_prevmap

DIFF_FILL = _PatternFill_for_prevmap(
    fill_type="solid", start_color="FFF2CC", end_color="FFF2CC"
)


# --- 差分ハイライト色（PREF.xlsx から可変） ------------------------------
import pandas as _pd_for_diffcolor

_DIFFF_DEFAULT_HEX = "FFF2CC"


def _hex6(s):
    if not s:
        return None
    s = str(s).strip()
    if s.startswith("#"):
        s = s[1:]
    return s.upper() if _re_for_prevmap.fullmatch(r"[0-9a-fA-F]{6}", s) else None


def load_diff_color(pref_source):
    """
    『色分け設定』シートから差分色を取得する。
    期待カラム: ENABLE / FIRST ROW / AND/OR / SECOND ROW / LABEL / COLOR
    優先: FIRST ROW == "__DIFF__" の行、無ければ LABEL に '差分' / '差分ハイライト' / 'diff'
    該当なしなら None（既定色を使う）。
    """
    try:
        df = _pd_for_diffcolor.read_excel(
            pref_source, sheet_name="色分け設定", dtype=str
        ).fillna("")
    except Exception:
        return None
    # 推定列名
    cols = {c.strip().lower(): c for c in df.columns}
    col_first = (
        cols.get("first row")
        or cols.get("first_row")
        or cols.get("first")
        or list(df.columns)[0]
    )
    col_label = cols.get("label") or cols.get("名称") or cols.get("説明")
    col_color = cols.get("color") or cols.get("色") or cols.get("カラー")
    col_enable = cols.get("enable") or cols.get("有効") or cols.get("enabled")

    def _enabled(r):
        if not col_enable:
            return True
        v = str(r.get(col_enable, "")).strip().lower()
        return v in ("yes", "true", "1", "y", "on", "")

    # 優先1: FIRST ROW == "__DIFF__"
    for _, r in df.iterrows():
        if not _enabled(r):
            continue
        fr = str(r.get(col_first, "")).strip()
        if fr == "__DIFF__":
            hx = _hex6(r.get(col_color, ""))
            if hx:
                return hx
    # 優先2: LABEL
    if col_label:
        for _, r in df.iterrows():
            if not _enabled(r):
                continue
            lb = str(r.get(col_label, "")).strip().lower()
            if any(k in lb for k in ("差分", "差分ﾊｲﾗｲﾄ", "diff")):
                hx = _hex6(r.get(col_color, ""))
                if hx:
                    return hx
    return None


# -------------------------------------------------------------------
def build_prev_map_from_excel(prev_xlsx_path):
    """
    前回出力Excelを (emp_no:str, day_index:int) -> value:str に変換。
    day_index はその人の「日付行で左から数えた列番号（1始まり）」。
    日付セルの表記（1/01/2025-08-01 等）の違いには影響されない。
    """
    wb = _load_wb_for_prevmap(prev_xlsx_path, data_only=True)
    ws = wb.active
    max_r, max_c = ws.max_row, ws.max_column

    def _is_dateish(v):
        if v is None:
            return False
        s = str(v).strip()
        return bool(
            _re_for_prevmap.match(r"^\d{1,2}$", s)
            or _re_for_prevmap.match(r"^\d{4}-\d{2}-\d{2}$", s)
        )

    emp_pat = _re_for_prevmap.compile(r"^(?:000)?(\d{5})$")

    # 1) 日付が複数並ぶ行を候補に
    date_rows = []
    for r in range(1, max_r + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_c + 1)]
        if sum(1 for v in vals if _is_dateish(v)) >= 5:
            date_rows.append(r)

    # 2) 各日付行の上～3行で職番を探す（000##### or 5桁）
    header_emp = {}  # row_of_dates -> emp_no(5桁)
    for r in date_rows:
        found = None
        for rr in range(max(1, r - 3), r):
            vals = [
                str(ws.cell(rr, c).value or "").strip() for c in range(1, max_c + 1)
            ]
            for v in vals:
                m = emp_pat.match(v)
                if m:
                    found = m.group(1)
                    break
            if found:
                break
        if found:
            header_emp[r] = found

    # 3) (emp_no, day_index) を構築（日付行の次行がスケジュール行）
    prev_map = {}
    for r in date_rows:
        emp_no = header_emp.get(r)
        if not emp_no:
            continue
        sched_row = r + 1
        date_cols = [c for c in range(1, max_c + 1) if _is_dateish(ws.cell(r, c).value)]
        for idx, c in enumerate(date_cols, start=1):  # idx: day_index
            val = str(ws.cell(sched_row, c).value or "").strip()
            prev_map[(emp_no, idx)] = val
    return prev_map


# ====================================================================

import warnings
import logging

# # ロガー設定（冒頭に記載）
# logging.basicConfig(
#     filename="simslot_debug.log",
#     level=logging.DEBUG,
#     format="%(asctime)s [%(levelname)s] %(message)s",
# )
# logger = logging.getLogger(__name__)

# ------------------------------------------
# 警告抑制設定
# ------------------------------------------
# openpyxl のスタイルシート関連の警告をまとめて無視
warnings.filterwarnings(
    "ignore", category=UserWarning, module="openpyxl.styles.stylesheet"
)

# ── Helpers: 補助関数群 ──────────────────────────────────────────────────

import os
import pandas as _pd
import io as _io


def load_sim_slot_excel(file):

    # ① file が None または空文字列なら即空 DataFrame
    if file is None or (isinstance(file, str) and file == ""):
        return _pd.DataFrame()

    try:
        if isinstance(file, str):
            if not os.path.exists(file):
                print(
                    f"[INFO] SIM Slot List ファイルが見つかりません ({file})。処理をスキップします。"
                )
                return _pd.DataFrame()
            df = _pd.read_excel(file, header=2, dtype=str)
        else:
            raw = file.read()
            df = _pd.read_excel(_io.BytesIO(raw), header=2, dtype=str)

        print(f"[INFO] SIM Slot: {len(df)} 行を読み込みました")
        return df

    except Exception as e:
        print(f"[ERROR] SIM Slot Excel 読み込み失敗: {e}。空の DataFrame を返します。")
        return _pd.DataFrame()


def load_simslot_schedule_codes(
    pref_file_path="PREF.xlsx", sheet_name="SIMSLOT"
) -> list:
    """
    PREF.xlsx の SIMSLOT シートから ActivityTypeCode のリストを取得する。

    - シート名が存在しない場合は空リストを返却
    - B1 セルにカンマ区切りでコードが設定されている想定

    Args:
        pref_file_path (str): 設定ファイルのパス
        sheet_name (str): 読み込むシート名 (デフォルト "SIMSLOT")

    Returns:
        list: 有効なコードのリスト
    """
    try:
        wb = load_workbook(pref_file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"✅ SIMSLOT シートが存在しません: {sheet_name}")
            return []

        ws = wb[sheet_name]
        b1_value = ws["B1"].value
        if not b1_value:
            print("✅ B1セルが空です")
            return []

        # カンマ区切りの文字列を分割し、前後空白を除去
        return [code.strip() for code in str(b1_value).split(",") if code.strip()]

    except Exception as e:
        print(f"⚠️ SIMSLOT 読み込み失敗: {e}")
        return []


def clean_cell(text: any) -> str:
    """
    セル内の文字列を正規化・クリーンアップする関数。

    処理内容:
      1. NaN / None を検出して空文字列に変換
      2. ゼロ幅スペースや NBSP、タブ、改行などの不可視文字を削除
      3. 文字列の前後の余分なスペースをトリム
      4. 最終的に空白のみの文字列は空文字列に統一

    Args:
        text: 任意の型で渡されるセルの元データ

    Returns:
        str: クリーンアップ後の文字列
    """
    import pandas as _pd
    import re as _re

    # 1) NaN / None を空文字に
    if _pd.isna(text):
        return ""
    s = str(text)
    # 2) 不可視文字や改行をすべて削除
    s = _re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s)
    # 3) 前後のスペースをトリム
    s = s.strip()
    # 4) 空文字か判定
    return s if s else ""


def remove_blank_and_ob(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame から不要な行を除去する関数。

    - 完全に空文字列の行を削除
    - 全セルが "" または "OB" のみの行を削除（Out-of-Bounds を想定）
    - 特定の Emp ID を含む行も除外可能（skip_ids を設定）

    Args:
        df: セルが文字列化され、クリーニング済みの pandas DataFrame

    Returns:
        pd.DataFrame: 空行と OB 行、特定 Emp ID 行を除去した DataFrame
    """
    # 除外する Emp ID のセット (例として3IDを設定)
    skip_ids = {"00049292", "00035957", "00041169"}

    rows = []
    for row in df.values:
        texts = [str(x).strip() for x in row]

        # 1) OB の既存ロジック: 99xxx 系や全 OB 行、末尾に OB が続く行を除去
        if any(re.fullmatch(r"00099[0-9]{3}", t) for t in texts):
            continue
        if all(not t or t == "OB" for t in texts):
            continue
        if re.fullmatch(r"[A-Z]+OB", texts[0]):
            continue

        # 2) 追加: 特定 Emp ID を含む行をスキップ
        if any(t in skip_ids for t in texts):
            continue

        rows.append(row)

    return pd.DataFrame(rows, columns=df.columns)


def find_header_rows(df: pd.DataFrame) -> list:
    """
    ヘッダー行を検出する関数。

    - 列0: 姓名+2レターが大文字英字で構成されるセル
    - 列2: 2レター (2文字の大文字英字)
    - 次行に日付行 (01～31の数値) が存在する

    Args:
        df (pd.DataFrame): クリーン済みの DataFrame

    Returns:
        list: ヘッダー行のインデックス一覧
    """
    hdrs = []
    for i in range(len(df) - 1):
        # 現行の1列目と3列目（0-indexed）をチェック
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        # 次行をリスト化して日付候補を確認
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]

        # 条件: c0 が 2文字以上大文字、c2 が 2文字大文字、次行に日付あり
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    return hdrs


def slice_blocks(df: pd.DataFrame) -> list:
    """
    ヘッダー／日付／スケジュール行ブロックを切り出す関数。

    ヘッダ行から次のヘッダ行までを1ブロックとして扱い、
    日付行位置とスケジュール列番号を算出する。

    Args:
        df (pd.DataFrame): クリーン済みの DataFrame

    Returns:
        list of tuples: (header_idx, date_idx, end_idx, date_columns)
    """
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)

    for idx, h in enumerate(hdrs):
        # 次のヘッダー行または末尾までを範囲とする
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        date_idx = h + 1
        # 最初の候補日付行が正しくなければ、次の行も探索
        vals = [str(df.iat[date_idx, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    date_idx = r
                    break
        # 日付列インデックス群を取得
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[date_idx, k]).strip() for k in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]

        blocks.append((h, date_idx, end, date_cols))
    return blocks


def load_pref_rules(file: str = "PREF.xlsx", sheet_name: str = "色分け設定") -> list:
    """
    色分けルールを Excel から読み込む。

    - シート2行目以降をルールとして取得
    - ENABLE が YES の行のみ有効
    - 6列目のセル背景色を取得し、カラーコードとして保存

    Args:
        file (str): 設定ファイルパス
        sheet_name (str): シート名

    Returns:
        list of dict: 各ルール辞書を返却
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] シート「{sheet_name}」が{file}に存在しません。")
        return []

    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=2, values_only=True))
    rules = []

    for idx, row in enumerate(data, start=2):
        enable, first, op, second, label, _ = row[:6]
        if str(enable).strip().upper() != "YES":
            continue
        # 6列目の背景色を取得
        cell = ws.cell(row=idx, column=6)
        color = (
            cell.fill.start_color.rgb
            if cell.fill
            and cell.fill.start_color
            and cell.fill.start_color.type == "rgb"
            else None
        )
        rules.append(
            {
                "first": str(first or "").strip(),
                "second": str(second or "").strip(),
                "op": str(op or "").strip().upper(),
                "label": str(label or "").strip(),
                "color": color or "",
            }
        )
    return rules


# ── 5) ルール適用ロジック ─────────────────────────────────────────────


def match_rule_in_multiline(
    text: any, rule: dict, debug_log_path: str = "debug_log.txt"
) -> bool:
    """
    複数行テキストに対し、PREFルールの条件1/条件2をAND/OR論理でチェックする。
    デバッグ情報はファイルに追記。

    Args:
        text (any): セル内の元テキスト（複数行想定）
        rule (dict): PREFから取得したルール辞書:
            {
              'first': str,  # 条件1
              'second': str, # 条件2
              'op': 'AND'|'OR'|'NONE'
            }
        debug_log_path (str): デバッグログ出力先パス
    Returns:
        bool: ルールにマッチすれば True
    """
    lines = str(text).split("\n")
    cond1 = rule.get("first", "")
    cond2 = rule.get("second", "")
    op = rule.get("op", "NONE").upper()

    cond1_found = any(cond1 in line for line in lines) if cond1 else False
    cond2_found = any(cond2 in line for line in lines) if cond2 else False

    # # デバッグ情報をログファイルに追記
    # with open(debug_log_path, "a", encoding="utf-8") as f:
    #     f.write("\n=== DEBUG MATCH ===\n")
    #     f.write(f"Text Lines: {lines}\n")
    #     f.write(f"Rule: {rule}\n")
    #     f.write(f"cond1_found={cond1_found}, cond2_found={cond2_found}, op={op}\n")

    if op == "AND":
        return cond1_found and cond2_found
    if op == "OR":
        return cond1_found or cond2_found
    # NONE なら条件いずれかでマッチ
    return cond1_found or cond2_found


def apply_pref_rules_to_cell(
    cell, val: any, rules: list, fallback_color: str = None
) -> None:
    """
    セルの文字列に対し、PREFルールを適用して背景色を設定する。

    Args:
        cell (openpyxl.cell): Excel セルオブジェクト
        val (any): セルに表示するスケジュール文字列
        rules (list): load_pref_rules で取得したルールリスト
        fallback_color (str): マッチしなかった場合に適用するデフォルトカラー
    """
    text = str(val).strip()
    lines = [line.strip() for line in text.split("\n") if line.strip()]

    for idx, rule in enumerate(rules):
        if not isinstance(rule, dict):
            print(f"⚠️ 無効なルール形式 (index={idx}): {rule}")
            continue

        first = rule.get("first", "")
        second = rule.get("second", "")
        op = rule.get("op", "NONE").upper()

        # 行ごとに正規表現マッチ
        cond1 = (
            any(re.search(first, line) for line in lines)
            if first not in (None, "")
            else False
        )
        cond2 = (
            any(re.search(second, line) for line in lines)
            if second not in (None, "")
            else False
        )

        if (
            (op == "AND" and cond1 and cond2)
            or (op == "OR" and (cond1 or cond2))
            or (op == "NONE" and (cond1 or cond2))
        ):
            color = rule.get("color", "").lstrip("#")
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            return

    # どのルールにもマッチしない場合、フォールバック色を設定
    if fallback_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fallback_color)


# ── 6) Excel 出力: 行ごとの書込み ─────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import re


import logging

# # ログレベルを DEBUG に設定（必要に応じて INFO 等に変更してください）
logging.basicConfig(level=logging.DEBUG)


def write_onboard_rows(
    ws,
    start_row: int,
    onboard_data: list[list[dict]],
    emp_aff_map: dict,
    block_aff: str,
    emp_to_row: dict,
    offset: int,
) -> int:
    """
    同乗クルー情報を Excel に行単位で書き込む（可変長＋ハイパーリンク＋オフセット対応）。
    デバッグ用にリンク先が取得できなかったケースをログに残します。

    Args:
        ws: openpyxl ワークシートオブジェクト
        start_row: 書込み開始行番号
        onboard_data: 各日付の同乗者情報リストのリスト
                      （要素は {"name": str, "emp_no": str or None} の dict）
        emp_aff_map: EmpNo→所属マップ
        block_aff: 現在のブロックの所属
        emp_to_row: EmpNo→ヘッダー行番号マップ
        offset: 本体データを何列右にずらすか（例：5）
    Returns:
        int: 書き込んだ同乗行数
    """
    # デバッグ用：リンク先が取れなかった同乗者を記録
    # missing = []

    # 各日付で最も同乗者数が多い行数を取得
    max_onb = max((len(day) for day in onboard_data), default=0)

    for i in range(max_onb):
        for j, day in enumerate(onboard_data, start=1):
            # 同乗者エントリ（存在しない場合は空 dict）
            entry = day[i] if i < len(day) else {}
            display = entry.get("name", "")
            emp_no = entry.get("emp_no")
            cell = ws.cell(row=start_row + i, column=j + offset)

            if display:
                # 社員番号に対応するヘッダー行番号を取得
                target = emp_to_row.get(emp_no)

                # デバッグ出力：同乗者ごとのマッピング状況を確認
                # logging.debug(
                #     f"[onboard i={i}, j={j}] display='{display}', emp_no='{emp_no}', target={target}"
                # )

                if emp_no and target:
                    # 正常ケース：表示名に社員番号を付与し、ハイパーリンクを設定
                    disp = f"{display} ({emp_no})"
                    link_col = get_column_letter(
                        offset + 1
                    )  # リンク先の列文字（F 列以降）
                    cell.value = f'=HYPERLINK("#Schedule!{link_col}{target}", "{disp}")'
                else:
                    # リンク先が取れない場合は単に表示名だけ
                    cell.value = display
                    # missing.append((display, emp_no, target))
            else:
                # 同乗者がいないセルは空文字
                cell.value = ""

            # 書式設定：左上揃え＋折り返し
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            # 同ブロック所属なら背景色ハイライト
            if emp_no and emp_aff_map.get(emp_no) == block_aff:
                cell.fill = PatternFill(fill_type="solid", fgColor="FFEE99")

    # ループ後、リンク先取得に失敗した同乗者をまとめてログ警告
    # if missing:
    #     logging.warning("リンク先取得失敗の同乗者一覧:")
    #     for disp, emp, tgt in missing:
    #         logging.warning(f"  display='{disp}', emp_no='{emp}', target={tgt}")

    return max_onb


from openpyxl.styles import Font


def write_to_excel(
    schedule_file: str,
    records: list,
    emp_aff_map: dict,
    out_xlsx: str,
    pref_rules: list,
    prev_map=None,
    diff_fill=None,
) -> None:
    """
    全レコードを Excel ファイルに書き出す（Ver31d 版）。

    ・先頭に5列追加（Index, Name, AI, AJ, Memo）
    ・ヘッダー～日付～スケジュール～同乗者行を順次配置
    ・同乗者情報は社員番号付き dict で保持し、ハイパーリンクを社員番号でマッピング
    """
    # 差分色のデフォルト設定（PREF未設定時は薄黄）
    if diff_fill is None:
        diff_fill = _PatternFill_for_prevmap(
            fill_type="solid", start_color="FFF2CC", end_color="FFF2CC"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # A～E 列を挿入してフィルター用ヘッダーを設定
    ws.insert_cols(1, amount=5)
    headers5 = ["Index", "Name", "AI", "AJ", "Memo"]
    for c, h in enumerate(headers5, start=1):
        ws.cell(row=1, column=c, value=h)
    offset = 5
    ws.freeze_panes = "A2"  # ← これだけで1行目固定！
    # 氏名→社員番号マップ（変更なし）
    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}

    # 社員番号→ヘッダー行番号マップ
    # --- generate_schedule31d.py
    # +++ generate_schedule31d.py
    # @@
    # -    emp_to_row: dict[str, int] = {}
    # -    r = 2  # データ開始行（ヘッダー行）が2行目なら2
    # -    for rec in records:
    # -        emp_no = rec["emp_no"]
    # -        # 現在の r が「この社員のヘッダー行」の行番号
    # -        emp_to_row[emp_no] = r
    # -
    # -        onb_cnt = sum(len(day) for day in rec.get("onb", []))
    # -        height = 3 + onb_cnt  # ヘッダー＋日付行＋スケジュール＋同乗者行
    # -        r += height
    # ■ 社員番号→ヘッダー行番号マップを作成
    emp_to_row: dict[str, int] = {}
    # ヘッダー行は 2 行目から開始（必要に応じて変更してください）
    current_row = 2

    for rec in records:
        emp_no = rec["emp_no"]

        # ① 当該レコードの同乗者データから、１日あたりの最大同乗者数を取得
        max_onboard = max((len(day) for day in rec.get("onb", [])), default=0)
        # ② ブロック全体の行数を計算（ヘッダー＋日付行＋スケジュール行＋同乗者行）
        block_height = 3 + max_onboard

        # ③ 現在の行番号を「この社員のヘッダー行番号」として登録
        emp_to_row[emp_no] = current_row

        # ④ 次ブロックの先頭行へカーソルを進める
        current_row += block_height

    # データを書き込む基準行
    current_row = 2
    for rec in records:
        # A～D 列（Index, Name, AI, AJ）
        emp_id = rec["emp_no"]
        hdr_alpha = rec["hdr"][0]
        raw_name = rec.get("orig_name", hdr_alpha[:-2])
        two = hdr_alpha[-2:]
        full_name = f"{raw_name} {two}"
        ai_val = rec["hdr"][29] if len(rec["hdr"]) > 29 else ""
        aj_val = rec["hdr"][30] if len(rec["hdr"]) > 30 else ""
        onb_cnt = sum(len(day) for day in rec.get("onb", []))
        block_h = 3 + onb_cnt

        for i in range(block_h):
            ws.cell(row=current_row + i, column=1, value=f"{emp_id}-{i+1}")
            ws.cell(row=current_row + i, column=2, value=full_name)
            ws.cell(row=current_row + i, column=3, value=ai_val)
            ws.cell(row=current_row + i, column=4, value=aj_val)

        # ヘッダー行（F 列以降）
        for ci, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=current_row, column=ci + offset, value=val)
            cell.border = Border(
                top=Side(border_style="double"), bottom=Side(border_style="double")
            )
            is_phone = bool(re.search(r"\d{2,4}-\d{2,4}-\d{4}", val)) or "電話" in val
            is_pe = bool(re.fullmatch(r"PE\d{6}", val))
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=not (is_phone or is_pe),
                shrink_to_fit=is_pe,
            )

        # ── ここから追記：E列＝養成期フィルター（数値のみ） ─────────────────
        # ラベルは1回だけ
        if ws.cell(row=1, column=5).value not in ("養成期",):
            ws.cell(row=1, column=5, value="養成期")

        # ✅ M列は rec["hdr"][7]（0-based）です。F=0, G=1, ..., M=7
        m_raw = str(rec["hdr"][7] if len(rec["hdr"]) > 7 else "").strip()
        phase_num = int(m_raw) if (m_raw.isdigit() and 2 <= len(m_raw) <= 3) else None

        ws.cell(
            row=current_row,
            column=5,
            value=phase_num if phase_num is not None else None,
        )
        # ────────────────────────────────────────────────────────────────

        # 日付行
        for ci, dv in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=current_row + 1, column=ci + offset, value=dv)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            apply_pref_rules_to_cell(
                cell,
                rec["sched"][ci - 1],
                pref_rules,
                fallback_color="DDDDDD",
            )

        # スケジュール行
        for ci, sv in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=current_row + 2, column=ci + offset, value=sv)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            # 差分ハイライト（新規/変更/削除すべて）
            if prev_map is not None:
                emp_id_str = str(emp_id)
                old = str(prev_map.get((emp_id_str, ci), ""))
                if str(sv or "") != old:
                    cell.fill = diff_fill
        # 同乗者情報を dict で組み立て → 書き込み
        raw_onb = rec.get("onb", [])
        onboard_data: list[list[dict]] = []
        for day in raw_onb:
            day_list: list[dict] = []
            for display in day:
                raw = display[1:] if display.startswith(("I", "T")) else display
                emp_no = name_to_emp.get(raw)
                day_list.append({"name": display, "emp_no": emp_no})
            onboard_data.append(day_list)

        onboard_count = write_onboard_rows(
            ws,
            current_row + 3,
            onboard_data,
            emp_aff_map,
            rec.get("aff"),
            emp_to_row,
            offset,
        )

        current_row += 3 + onboard_count

    # フィルター範囲を A～E に（E=養成期フィルター列）
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    # ② 可変列数対応の CSV 読み込み（Streamlit BytesIO／ファイルパス 両対応）
    import csv, io, os

    # ── 生テキスト取得 ────────────────────────────────────────
    if isinstance(schedule_file, (str, os.PathLike)):
        with open(schedule_file, "r", encoding="utf-8", newline="") as fp:
            raw_text = fp.read()
    elif hasattr(schedule_file, "read"):
        schedule_file.seek(0)
        raw = schedule_file.read()
        raw_text = raw.decode("utf-8") if isinstance(raw, (bytes, bytearray)) else raw
    else:
        raise ValueError(f"Unsupported schedule_file type: {type(schedule_file)}")
    # ────────────────────────────────────────────────────────

    # ── 行単位に分割して reader へ ─────────────────────────────
    lines = raw_text.splitlines()
    reader = csv.reader(lines)
    rows = list(reader)
    if not rows:
        raise ValueError("Schedule CSV が空か、解析できませんでした。")
    # ────────────────────────────────────────────────────────

    # ── 最大列数でパディング ────────────────────────────────────
    max_cols = max(len(r) for r in rows)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]
    # ────────────────────────────────────────────────────────

    # ── pandas DataFrame 化 ────────────────────────────────────
    sched_df = pd.DataFrame(padded, dtype=str).fillna("")
    first_row = sched_df.iloc[0].astype(str).tolist()
    # ────────────────────────────────────────────────────────────

    # ③ 第1行から YYYYMM と FLEET を抽出
    month = ""
    fleet = ""
    for cell in first_row:
        c = cell.strip().lstrip("\ufeff")
        if not month:
            m = re.search(r"(\d{6})", c)
            if m:
                month = m.group(1)
        if not fleet:
            m2 = re.search(r"FLEET[:：]\[(.+?)\]", c, flags=re.IGNORECASE)
            if m2:
                fleet = m2.group(1)
        if month and fleet:
            break

    ws["F1"] = f"{fleet} 乗員スケジュール"
    ws["F1"].font = Font(size=20, bold=True)

    ws["J1"] = month
    ws["J1"].font = Font(size=16)
    # ── 追加フォーマット設定 ──────────────────────────────────────────
    # A列を非表示
    ws.column_dimensions["A"].hidden = True

    # B〜E列を幅3に
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 3

    # A〜E列の全セルを白フォントに
    # ※書き込むデータ数に合わせて max_row は後で上書きしてもOK
    #    （データ書き込み後に再度ループしても構いません）
    for col in ("A", "B", "C", "D", "E"):
        for row in range(1, ws.max_row + 1):
            ws[f"{col}{row}"].font = Font(color="FFFFFF")

    # シート全体のデフォルト行高を20に（これを先頭で）
    # ws.sheet_format.defaultRowHeight = 20
    # そして 1行目だけ改めて高さ指定
    ws.row_dimensions[1].height = 20

    # --- Legend（凡例）シート作成（差分色のみ） --------------------------
    try:
        _wb = ws.parent  # Workbook
        # 既存Legendがあれば作り直し
        if "Legend" in [s.title for s in _wb.worksheets]:
            _wb.remove(_wb["Legend"])
        ws_legend = _wb.create_sheet("Legend")
        ws_legend.append(["説明", "色見本"])
        demo = ws_legend.cell(row=2, column=2, value="差分セル")
        demo.fill = diff_fill
        ws_legend.cell(
            row=2, column=1, value="前回Excelと今回の値が異なるセルをハイライト"
        )
        ws_legend.freeze_panes = "A2"
        try:
            ws_legend.column_dimensions["A"].width = 48
            ws_legend.column_dimensions["B"].width = 14
        except Exception:
            pass
    except Exception:
        pass
    # -------------------------------------------------------------------

    ws.parent.save(out_xlsx)


import pandas as pd
import re
import csv
import os
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def run(
    schedule_file,
    pref_file,
    sim_file,
    prev_excel=None,
) -> tuple[str, str]:
    # ── ファイルポインタを先頭に戻す ───────────────────────────
    for f in (schedule_file, pref_file, sim_file):
        if hasattr(f, "seek"):
            f.seek(0)
    # ────────────────────────────────────────────────────

    # スケジュール CSV 読み込み（可変列数＆BytesIO対応）
    import csv, io, os

    # ── 生テキストを取得 ───────────────────────────────────
    if isinstance(schedule_file, (str, os.PathLike)):
        with open(schedule_file, "r", encoding="utf-8", newline="") as fp:
            raw_text = fp.read()
    elif hasattr(schedule_file, "read"):
        schedule_file.seek(0)
        raw = schedule_file.read()
        raw_text = raw.decode("utf-8") if isinstance(raw, (bytes, bytearray)) else raw
    else:
        raise ValueError(f"Unsupported schedule_file type: {type(schedule_file)}")
    # ─────────────────────────────────────────────────────

    # ── 行単位に分割して reader へ ────────────────────────
    lines = raw_text.splitlines()
    reader = csv.reader(lines)
    rows = list(reader)
    if not rows:
        raise ValueError("Schedule CSV が空か、解析できませんでした。")
    # ─────────────────────────────────────────────────────

    # ── 最大列数でパディング ───────────────────────────────
    max_cols = max(len(r) for r in rows)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]
    # ─────────────────────────────────────────────────────

    # ── pandas DataFrame 化 ─────────────────────────────────
    sched = pd.DataFrame(padded, dtype=str).fillna("")
    # ─────────────────────────────────────────────────────

    # PREF ファイル読み込み（str or BytesIO 対応）
    pref = (
        pref_file
        if isinstance(pref_file, (str, os.PathLike))
        else io.BytesIO(pref_file.read())
    )
    emp_df = pd.read_excel(pref, sheet_name="emp_no", header=None, dtype=str).fillna("")

    # 色分けルールと SIM Slot コード読み込み（補助関数が外にある想定）
    pref_rules = load_pref_rules(pref)
    simslot_codes = load_simslot_schedule_codes(pref)

    # # SIM Slot ファイル読み込み（str or BytesIO 対応）
    # sim = (
    #     sim_file
    #     if isinstance(sim_file, (str, os.PathLike))
    #     else io.BytesIO(sim_file.read())
    # )
    # wb_sim = load_workbook(sim, read_only=True, data_only=True)
    # first_sheet = wb_sim.sheetnames[0]
    # sim_df = pd.read_excel(sim, sheet_name=first_sheet, header=2, dtype=str).fillna("")

    # ── 2) 社員情報マップ作成 ─────────────────────────────────────────
    # emp_df から各種マップを生成: 氏名、2レター、所属、背景設定列
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()
    # 養成期 (D 列) と Phase (H 列) のマッピング
    phase_d_map = {
        str(row[2]).strip(): str(row[3]).strip() for _, row in emp_df.iterrows()
    }
    phase_h_map = {
        str(row[2]).strip(): str(row[7]).strip() for _, row in emp_df.iterrows()
    }

    # ── 3) 元データ整形 ────────────────────────────────────────────
    # セルクリーンと空行・OB 行削除
    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)

    # OBなどヘッダーだけでスケジュールが存在しない行を除外（NEW）
    def remove_ob_only_headers(df):
        def is_valid_header(idx):
            if idx + 1 >= len(df):  # 次行が存在しない
                return False
            next_row = df.iloc[idx + 1]
            # 次の行に日付らしき01〜31が5つ以上あるか
            has_date = (
                next_row.astype(str).str.match(r"^(0[1-9]|[12][0-9]|3[01])$").sum() >= 5
            )
            return has_date

        # ヘッダーだけの行や列数が極端に少ない行を除外
        mask = [is_valid_header(i) or df.iloc[i].count() > 30 for i in range(len(df))]
        return df[mask]

    # ↑ 上記のフィルタ関数を適用
    df = remove_ob_only_headers(df)

    # ブロック抽出
    blocks = slice_blocks(df)
    if not blocks:
        return
    global_dates = blocks[0][3]

    # ── 4) SIM Slot 実績読み込み & 参加者辞書作成 ────────────────────────
    from datetime import datetime

    # SIM Slot 読み込み（ファイル未提供 or 読み込み失敗時も空 DataFrame → 空辞書でフォールバック）
    sim_df = load_sim_slot_excel(sim_file)

    if sim_df.empty:
        # SIM Slot がない場合は空辞書を用意
        teacher_lookup = {}
        trainee_lookup = {}
        simslot_participants = {}
    else:
        # ── SIM Slot 処理ロジック開始 ────────────────────────────────

        # '日付' 列を 'day' に統一
        if "日付" in sim_df.columns:
            sim_df.rename(columns={"日付": "day"}, inplace=True)
        else:
            sim_df["day"] = pd.NA

        # Event Name 列を確保
        sim_df["Event Name"] = sim_df.get("Event Name", "").astype(str).str.strip()

        # 'day' 列を Int64 に変換
        try:
            sim_df["day"] = pd.to_datetime(
                sim_df["day"], errors="coerce"
            ).dt.day.astype("Int64")
        except Exception:
            pass

        # 教官／訓練生 ID 列を文字列化→欠損→分割して必ず list 化
        sim_df["教官 Emp ID"] = (
            sim_df.get("教官 Emp ID", "")
            .fillna("")  # NaN → ""
            .astype(str)  # float → "nan" ではなく元を文字列化
            .str.split("/", expand=False)
            .apply(lambda lst: [x for x in lst if x and x.strip()])  # 空要素除去
        )
        sim_df["訓練生 Emp ID"] = (
            sim_df.get("訓練生 Emp ID", "")
            .fillna("")
            .astype(str)
            .str.split("/", expand=False)
            .apply(lambda lst: [x for x in lst if x and x.strip()])
        )

        # ActivityTypeCode でフィルタ
        sim_df = sim_df[sim_df["ActivityTypeCode"].isin(simslot_codes)]

        # 号機列はシート2列目 (index=1) を使用
        machine_col = sim_df.columns[1]

        # A350/787 号機判定用ユーティリティ
        def get_machine(raw):
            if pd.isna(raw) or str(raw).strip() == "":
                return "APT"
            raw = str(raw).strip()
            if raw in {"1", "2", "7", "8"}:
                return f"#{raw}"
            return raw

        # 参加者辞書初期化
        teacher_lookup = {}
        trainee_lookup = {}
        simslot_participants = {}

        # 各行をループして辞書を構築
        for _, row in sim_df.iterrows():
            day = row.get("day")
            act_code = row.get("ActivityTypeCode", "").strip()
            evt_code = row.get("Event Name", "").strip()
            start = row.get("開始時刻", "")
            end = row.get("終了時刻", "")
            machine = get_machine(row.get(machine_col, ""))

            teachers = [eid.strip()[-5:] for eid in row["教官 Emp ID"] if eid.strip()]
            trainees = [eid.strip()[-5:] for eid in row["訓練生 Emp ID"] if eid.strip()]

            # 時間帯登録
            for eid in teachers:
                teacher_lookup[(day, eid)] = (start, end)
            for eid in trainees:
                trainee_lookup[(day, eid)] = (start, end)

            # 参加者辞書キー (act_code ＋ evt_code)
            codes = [act_code] + (
                [evt_code] if evt_code and evt_code != act_code else []
            )
            for code_key in codes:
                key = (day, code_key, start, end, machine)
                simslot_participants.setdefault(
                    key, {"teachers": [], "trainees": [], "event": evt_code}
                )
                simslot_participants[key]["teachers"].extend(teachers)
                simslot_participants[key]["trainees"].extend(trainees)

        # 重複除去
        for grp in simslot_participants.values():
            grp["teachers"] = list(dict.fromkeys(grp["teachers"]))

        # ── SIM Slot 処理ロジック終了 ────────────────────────────────

    # ── 5) records 作成 ─────────────────────────────────────────────────
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000\d{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        # 元のヘッダー（MATSUNAMI など）を保持
        orig_name = raw[0]

        raw[0] = f"{surname}{two}" if matched else raw[0]

        hdr_vals = [v for v in raw if v]
        hdr = hdr_vals[:31] + [""] * (31 - len(hdr_vals[:31]))

        # ★追加：このクルーのスケジュール本文に「VLD」が含まれるかをブロック内で走査
        #        （日付行 d の次行以降～ブロック終端 end までをざっくりチェック）
        vld_in_sched = False
        try:
            for rr in range(d + 1, end + 1):
                row_vals = [str(clean_cell(x)) for x in df.iloc[rr]]
                if any("VLD" in v for v in row_vals):
                    vld_in_sched = True
                    break
        except Exception:
            vld_in_sched = False

        # 養成期/Phase 表示を設定
        raw_d = phase_d_map.get(code, "").strip()
        raw_h = phase_h_map.get(code, "").strip()

        # ★変更：VLDが含まれ、かつ D列が「数字のみ」の場合は CMDV を強制表示
        if vld_in_sched and raw_d.isdigit():
            display = "CMDV"

        elif raw_d and not raw_d.isdigit():
            # D列がアルファベット系のとき
            if raw_d == "B":
                display = "B教官"
            elif raw_d == "BS":
                display = "S教官"
            elif raw_d == "BSL":
                display = "L教官"
            elif raw_d == "VBSL":
                display = "VLD"
            else:
                display = raw_d
        else:
            # D列が数字のみ or 空文字 → H列を参照（従来通り）
            if raw_h.isdigit():
                display = f"PH{raw_h}"
            else:
                display = raw_h

        # hdr[29] に表示文字列を格納（ご指定通り）
        hdr[29] = display
        hdr[30] = rec_aff

        # ── セル内文字列整形ルール (まとめて置換) ───────────────────────────────────
        # 「PE有効期限」→「PE」, 「PExxxxxx」→「xxxxxx」
        # 「社員番号」→「職番」, 「電話番号」→「電話」
        # 「CAT資格」→「CAT資」, 「T/O期限」→「T/O」, 「L/D期限」→「L/D」
        hdr = [
            re.sub(
                r"L/D期限",
                "L/D",
                re.sub(
                    r"T/O期限",
                    "T/O",
                    re.sub(
                        r"CAT資格",
                        "CAT資",
                        re.sub(
                            r"電話番号",
                            "電話",
                            re.sub(
                                r"社員番号",
                                "職番",
                                re.sub(
                                    r"PE(\\d{6})",
                                    r"\\1",
                                    re.sub(r"PE有効期限", "PE", v),
                                ),
                            ),
                        ),
                    ),
                ),
            )
            for v in hdr
        ]

        # 日付行・スケジュール行の取得（既存ロジック）
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            fe.append(
                [
                    clean_cell(df.iat[r2, j])
                    for r2 in range(d + 1, end)
                    if clean_cell(df.iat[r2, j])
                ]
            )
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))

        # レコードに追加
        records.append(
            {
                "emp_no": code,
                "orig_name": orig_name,  # ★ ここを追加 ★
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )

    # 全日数埋め
    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))

    # ── 6) Phase1: 訓練コード＋ISR/TRN のみ残す ──────────────────────────────
    for rec in records:
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            first = cell_text.split("\n", 1)[0].strip()
            if first in simslot_codes:
                if "ISR" in cell_text:
                    rec["sched"][idx] = f"{first}\nISR"
                elif "TRN" in cell_text:
                    rec["sched"][idx] = f"{first}\nTRN"

    # ── 7) Phase2: 時間帯＋号機＋Event Name（特定コードのみ）追記 ─────────
    for rec in records:
        emp_id = rec["emp_no"]
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            lines = cell_text.split("\n")
            code = lines[0].strip()
            if code not in simslot_codes or len(lines) < 2:
                continue
            subcode = lines[1].strip()
            if subcode not in ("ISR", "TRN"):
                continue

            # 日付取得
            try:
                day = int(rec["dr"][idx])
            except ValueError:
                continue

            # 時刻取得
            lookup = teacher_lookup if subcode == "ISR" else trainee_lookup
            time_pair = lookup.get((day, emp_id))
            if not time_pair:
                continue
            start, end = time_pair

            # 号機特定
            machine = None
            parts = None
            for (d, c, s, e, m), grp in simslot_participants.items():
                if (d, c, s, e) == (day, code, start, end) and emp_id in grp[
                    "teachers"
                ] + grp["trainees"]:
                    machine = m
                    parts = grp
                    break
            if not parts:
                continue

            # Event Name は特定コードのみ
            event_name = parts.get("event", "")
            base = f"{code}\n{subcode}\n{machine}\n{start}-{end}"
            if code in ("FOTR", "CATR", "CAUG") and event_name:
                rec["sched"][idx] = f"{base}\n{event_name}"
            else:
                rec["sched"][idx] = base
    # ───────────────────────────────────────────────────────────────────

    # ── 8) Phase3: onb ロジック統合（自己除外強化版） ────────────────────────
    for i, rec in enumerate(records):
        emp_id = rec["emp_no"]
        onb = []
        for idx, sched_cell in enumerate(rec["sched"]):
            if not sched_cell:
                onb.append([])
                continue
            lines = sched_cell.split("\n", 1)
            first = lines[0].strip()

            if first in simslot_codes:
                try:
                    day = int(rec["dr"][idx])
                except ValueError:
                    onb.append([])
                    continue

                parts = None
                machine = None

                # Debugログ：マッチング試行を記録
                # logging.debug(
                #     f"🔍 SIM マッチチェック: day={day}, code={first}, emp_id={emp_id}"
                # )

                for (d, c, s, e, m), grp in simslot_participants.items():
                    if (
                        d == day
                        and c == first
                        and emp_id in grp["teachers"] + grp["trainees"]
                    ):
                        parts = grp
                        machine = m
                        # logging.debug(
                        #     f"✅ MATCH: key=({d}, {c}, {s}, {e}, {m}), emp_id={emp_id}"
                        # )
                        break
                    # else:
                    #     logging.debug(
                    #         f"… NO match: key=({d}, {c}, {s}, {e}, {m}), emp_id={emp_id} not in {grp['teachers'] + grp['trainees']}"
                    #     )

                if not parts:
                    # logging.debug(
                    #     f"❌ NO MATCH FOUND for emp_id={emp_id}, code={first}, day={day}"
                    # )
                    onb.append([])
                    continue

                names = []
                for tid in parts["teachers"]:
                    if tid.strip() == emp_id.strip():
                        continue
                    nm = emp_name_map.get(tid, "")
                    two = emp_two_map.get(tid, "")
                    if nm:
                        names.append(f"I{nm}{two}")
                for tid in parts["trainees"]:
                    if tid.strip() == emp_id.strip():
                        continue
                    nm = emp_name_map.get(tid, "")
                    two = emp_two_map.get(tid, "")
                    if nm:
                        names.append(f"T{nm}{two}")
                onb.append(names)
            else:
                entries = rec["full_entries"][idx]
                flights = [e for e in entries if e and re.match(r"^[0-9]", e)]
                names = []
                for j, other in enumerate(records):
                    if i == j or other["hdr"][0] == rec["hdr"][0]:
                        continue
                    other_entries = other["full_entries"][idx]
                    if any(f in other_entries for f in flights):
                        names.append(other["hdr"][0])
                uniq = []
                for n in names:
                    if n not in uniq:
                        uniq.append(n)
                onb.append(uniq)
        rec["onb"] = onb

    # ── 9) 重複レコード削除＆ソート ────────────────────────────────────────────
    seen, uniq = set(), []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = sorted(
        uniq,
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        ),
    )

    from datetime import datetime

    # → 既存の「today = …」「base = …」あたりを丸ごと置き換え
    # MMDDhhmm（2桁月・2桁日・2桁時・2桁分）形式のタイムスタンプを取得
    timestamp = datetime.now().strftime("%m%d%H%M")

    # NAGU＋タイムスタンプ の形式で出力
    out_csv = f"NAGU{timestamp}.csv"
    out_xlsx = f"NAGU{timestamp}.xlsx"

    # CSV 出力
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec.get("onb", [])])

    # 差分色を PREF.xlsx から取得（色分け設定の __DIFF__ 行など）
    try:
        _diff_hex = load_diff_color(pref_file) or _DIFFF_DEFAULT_HEX
    except Exception:
        _diff_hex = _DIFFF_DEFAULT_HEX
    diff_fill = _PatternFill_for_prevmap(
        fill_type="solid", start_color=_diff_hex, end_color=_diff_hex
    )
    # Excel 出力前に、前回Excelが渡されていれば比較マップを構築
    prev_map = None
    if prev_excel:
        import io as _io_prev, tempfile as _tmp_prev

        if isinstance(prev_excel, (bytes, bytearray, _io_prev.BytesIO)):
            data = (
                prev_excel
                if isinstance(prev_excel, (bytes, bytearray))
                else prev_excel.getvalue()
            )
            tmp = _tmp_prev.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.write(data)
            tmp.flush()
            tmp.close()
            prev_path = tmp.name
        else:
            prev_path = str(prev_excel)
        try:
            prev_map = build_prev_map_from_excel(prev_path)
        except Exception:
            prev_map = None

    # Excel 出力
    write_to_excel(
        schedule_file,
        records,
        emp_aff_map,
        out_xlsx,
        pref_rules,
        prev_map=prev_map,
        diff_fill=diff_fill,
    )

    return out_csv, out_xlsx


import argparse


def main():
    import argparse
    from pathlib import Path

    p = argparse.ArgumentParser(
        description="クルースケジュール生成ツール（ローカルテスト用）"
    )
    p.add_argument(
        "--mode",
        choices=("350", "787"),
        default="350",
        help="使用データを選択（350＝A350用、787＝B787用）",
    )
    # 任意指定（指定があれば mode の既定より優先）
    p.add_argument("--schedule", help="スケジュールCSVのパス（任意指定）")
    p.add_argument("--pref", help="PREF.xlsx のパス（任意指定）")
    p.add_argument(
        "--sim", help="SIM Slot List のパス（任意指定。空にしたい場合は --sim ''）"
    )
    p.add_argument(
        "--prev", help="前回の出力Excel(.xlsx) のパス（比較ハイライトしたい場合）"
    )
    args = p.parse_args()

    # スクリプト配置ディレクトリを基準に相対パスを解決
    BASE = Path(__file__).resolve().parent

    def resolve(p):
        """相対パスをこのスクリプト基準で絶対化。空文字/Noneはそのまま返す。"""
        if p is None:
            return None
        s = str(p)
        if s == "":
            return ""
        pp = Path(s)
        return str(pp if pp.is_absolute() else (BASE / pp))

    # 既定セット（必要に応じてここを書き換えてOK）
    if args.mode == "787":
        schedule_file = args.schedule or "schedule_787.csv"
        pref_file = args.pref or "PREF_787_harada.xlsx"
        sim_file = args.sim or "SIM Slot List 202507_787.xlsx"
    else:
        schedule_file = args.schedule or "schedule350_08.csv"
        pref_file = args.pref or "PREF350.xlsx"
        # ⚠️ ファイル名の拡張子位置を修正
        sim_file = args.sim if args.sim is not None else "202508 SIM Slot List.xlsx"
        # sim_file を使わない場合は --sim "" と指定してください

    # 絶対パスへ解決
    schedule_file = resolve(schedule_file)
    pref_file = resolve(pref_file)
    sim_file = resolve(sim_file) if sim_file is not None else ""
    prev_excel = resolve(args.prev) if args.prev else None

    # 実行
    csv_path, xlsx_path = run(
        schedule_file=schedule_file,
        pref_file=pref_file,
        sim_file=sim_file,
        prev_excel=prev_excel,  # ← 前回Excelを渡すとハイライト動作
    )

    print(f"[DONE] CSV: {csv_path}")
    print(f"[DONE] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
