# === generate_schedule.py ===
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ====


def clean_cell(text):
    """Remove invisible characters and trim whitespace"""
    s = str(text)
    return re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s).strip()


def remove_blank_and_ob(df):
    """Drop rows that are all blank or contain only 'OB'"""
    df = df.replace({pd.NA: "", "nan": "", "NaN": ""})
    df = df[(df != "").any(axis=1)]
    df = df[~df.apply(lambda row: all(str(c).strip() == "OB" for c in row), axis=1)]
    return df.reset_index(drop=True)


def find_header_rows(df):
    """
    Detect header rows where:
    - first cell ends with two uppercase letters
    - next row has exactly 31 date values (1-31)
    """
    hdrs = []
    for i in range(len(df) - 1):
        first = clean_cell(df.iat[i, 0])
        if re.search(r"[A-Z]{2}$", first):
            next_row = [clean_cell(x) for x in df.iloc[i + 1]]
            dates = [
                v for v in next_row if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
            ]
            if len(dates) == 31:
                hdrs.append(i)
    return hdrs


def slice_blocks(df):
    """Slice into blocks by header and first date row"""
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        for r in range(h + 1, end):
            row = [clean_cell(x) for x in df.iloc[r]]
            dates = [
                j
                for j, v in enumerate(row)
                if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
            ]
            if dates:
                blocks.append((h, r, end, dates))
                break
    return blocks


def write_to_excel(records, out_xlsx):
    wb = Workbook()
    ws = wb.active
    grey = PatternFill(fill_type="solid", fgColor="DDDDDD")
    hi = PatternFill(fill_type="solid", fgColor="FFFF00")
    dbl = Side(border_style="double", color="000000")
    brd = Border(top=dbl, bottom=dbl)

    for c in range(1, 32):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 8

    r = 1
    for rec in records:
        # Header
        for c, v in enumerate(rec["hdr"], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.border = brd
        r += 1
        # Dates
        for c, v in enumerate(rec["dr"], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = grey
        r += 1
        # Schedule
        for c, v in enumerate(rec["sched"], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        r += 1
        # Onboard
        for c, v in enumerate(rec["onb"], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if "*" in v:
                cell.fill = hi
        r += 1
    wb.save(out_xlsx)


def run(schedule_input, emp_input, config_path=None):
    # Load
    sched = pd.read_csv(schedule_input, header=None, dtype=str).fillna("")
    emp = pd.read_csv(emp_input, header=None, dtype=str).fillna("")
    # Emp map
    emp_map = {}
    for _, row in emp.iterrows():
        code = str(row[2]).zfill(5)
        if re.fullmatch(r"[0-9]{5}", code):
            emp_map[code] = (row[4] or "") + (row[6] or "")
    # Clean
    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        # map header
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        if matched:
            c = matched[0][3:]
            if c in emp_map:
                raw[0] = emp_map[c]
        # format hdr
        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        # dates
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        # sched
        fe = []
        for j in dates:
            es = []
            for r2 in range(d + 1, end):
                v = clean_cell(df.iat[r2, j])
                if v:
                    es.append(v)
            fe.append(es)
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))
        records.append({"hdr": hdr, "dr": dr, "sched": sched_row, "full_entries": fe})
    # onboard
    for i, cd in enumerate(records):
        onb = []
        for idx, es in enumerate(cd["full_entries"]):
            flights = [e for e in es if re.match(r"^[0-9]", e)]
            names = []
            for j, other in enumerate(records):
                if i == j:
                    continue
                of = [e for e in other["full_entries"][idx] if re.match(r"^[0-9]", e)]
                if any(f in of for f in flights):
                    names.append(other["hdr"][0])
            onb.append("\n".join(names))
        onb += [""] * (31 - len(onb))
        cd["onb"] = onb
    # output
    out_csv = "formatted_schedule.csv"
    rows = []
    for r in records:
        rows.extend([r["hdr"], r["dr"]] + r["sched"] + [r["onb"]])
    pd.DataFrame(rows).to_csv(out_csv, index=False, header=False)
    out_xlsx = "formatted_schedule.xlsx"
    write_to_excel(records, out_xlsx)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--emp", default="emp_no.csv")
    a = p.parse_args()
    run(a.schedule, a.emp)
