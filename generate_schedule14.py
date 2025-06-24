# === generate_schedule.py ===
import pandas as pd
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ====


def clean_cell(text):
    """Remove invisible characters and trim whitespace"""
    s = str(text)
    return re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s).strip()


def remove_blank_and_ob(df):
    """Drop rows that are all blank or contain only 'OB'"""
    clean_rows = []
    for row in df.values:
        texts = [str(x).strip() for x in row]
        if all(not t or t == "OB" for t in texts):
            continue
        clean_rows.append(row)
    return pd.DataFrame(clean_rows, columns=df.columns)


def find_header_rows(df):
    """Identify header rows by name and two-letter code patterns"""
    hdrs = []
    for i in range(len(df) - 1):
        first = str(df.iat[i, 0]).strip()
        second = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        if re.fullmatch(r"[A-Z]{2,}", first) and re.fullmatch(r"[A-Z]{2}", second):
            # next row has at least one date value
            next_vals = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]
            if any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_vals if v
            ):
                hdrs.append(i)
    return hdrs


def slice_blocks(df):
    """Slice into header/date blocks"""
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        d = h + 1
        # ensure d is date row, else fallback
        date_vals = [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
        if not any(
            re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in date_vals if v
        ):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    d = r
                    break
        # collect date columns
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]
        blocks.append((h, d, end, date_cols))
    return blocks


def write_to_excel(records, out_xlsx):
    wb = Workbook()
    ws = wb.active
    grey = PatternFill(fill_type="solid", fgColor="DDDDDD")
    double = Side(border_style="double", color="000000")

    row_num = 1
    for rec in records:
        # header
        for j, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.border = Border(top=double, bottom=double)
        # dates
        for j, val in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=row_num + 1, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = grey
        # schedule
        for j, val in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num + 2, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        # onboard
        for j, names in enumerate(rec["onb"], start=1):
            cell = ws.cell(row=row_num + 3, column=j, value="\n".join(names))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        row_num += 4
    wb.save(out_xlsx)


def run(schedule_file, emp_file):
    sched = pd.read_csv(schedule_file, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_file, header=None, dtype=str).fillna("")

    # maps for surname and two-letter
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}

    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks_raw = slice_blocks(df)
    # unify date columns from first block
    if blocks_raw:
        _, _, _, global_dates = blocks_raw[0]
    else:
        global_dates = []
    # rebuild blocks with same date columns
    blocks = [(h, d, end, global_dates) for (h, d, end, _) in blocks_raw]

    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        if matched:
            code = matched[0][3:]
            surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
            two_letter = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
            raw[0] = f"{surname}{two_letter}"
        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            entries = [
                clean_cell(df.iat[r2, j])
                for r2 in range(d + 1, end)
                if clean_cell(df.iat[r2, j])
            ]
            fe.append(entries)
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))
        records.append({"hdr": hdr, "dr": dr, "sched": sched_row, "full_entries": fe})

    # compute onboard lists
    for rec in records:
        onb = []
        for idx in range(len(global_dates)):
            entries = rec["full_entries"][idx]
            flights = [e for e in entries if re.match(r"^[0-9]", e)]
            names = []
            for other in records:
                if other is rec:
                    continue
                other_entries = other["full_entries"][idx]
                other_flights = [e for e in other_entries if re.match(r"^[0-9]", e)]
                if any(f in other_flights for f in flights):
                    names.append(other["hdr"][0])
            onb.append(names)
        rec["onb"] = onb

    # CSV output (4 lines per record)
    out_csv = "formatted_schedule.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for rec in records:
            writer.writerow(rec["hdr"])
            writer.writerow(rec["dr"])
            writer.writerow(rec["sched"])
            writer.writerow(["\n".join(lst) for lst in rec["onb"]])

    # Excel output
    out_xlsx = "formatted_schedule.xlsx"
    write_to_excel(records, out_xlsx)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--schedule", default="schedule.csv")
    parser.add_argument("--emp", default="emp_no.csv")
    args = parser.parse_args()
    run(args.schedule, args.emp)
