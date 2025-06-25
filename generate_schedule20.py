#!/usr/bin/env python3
# === generate_schedule19.py ===

import pandas as pd
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ===


def clean_cell(text):
    """Remove invisible characters and trim whitespace."""
    return re.sub(
        r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", str(text)
    ).strip()


def remove_blank_and_ob(df):
    """Drop rows that are all blank or contain only 'OB'."""
    rows = []
    for row in df.values:
        texts = [str(x).strip() for x in row]
        if all(not t or t == "OB" for t in texts):
            continue
        rows.append(row)
    return pd.DataFrame(rows, columns=df.columns)


def find_header_rows(df):
    """Identify header rows by name and two-letter code patterns."""
    hdrs = []
    for i in range(len(df) - 1):
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    return hdrs


def slice_blocks(df):
    """Slice into blocks: (header_idx, date_idx, end_idx, date_cols)."""
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        d = h + 1
        vals = [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    d = r
                    break
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]
        blocks.append((h, d, end, date_cols))
    return blocks


# Styles for Excel
HIGHLIGHT = PatternFill(fill_type="solid", fgColor="FFEE99")
DEFAULT_GREY = PatternFill(fill_type="solid", fgColor="DDDDDD")
DOUBLE = Side(border_style="double", color="000000")


def write_to_excel(records, emp_aff_map, out_xlsx):
    wb = Workbook()
    ws = wb.active
    row_num = 1
    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}

    for rec in records:
        block_aff = rec["aff"]
        # header row
        for j, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=j, value=val)
            wrap = not bool(re.fullmatch(r"0[0-9]{1,}-[0-9]+-[0-9]{4}", val))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=wrap
            )
            cell.border = Border(top=DOUBLE, bottom=DOUBLE)
        # date row
        for j, val in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=row_num + 1, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = DEFAULT_GREY
        # schedule row
        for j, val in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num + 2, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        # onboard row: highlight cells with same affiliation
        for j, names in enumerate(rec["onb"], start=1):
            cell = ws.cell(row=row_num + 3, column=j, value="\n".join(names))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            for name in names:
                emp = name_to_emp.get(name)
                if emp and emp_aff_map.get(emp) == block_aff:
                    cell.fill = HIGHLIGHT
                    break
        row_num += 4
    wb.save(out_xlsx)


# ==== Main Logic ===
def run(schedule_file, emp_file):
    sched = pd.read_csv(schedule_file, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_file, header=None, dtype=str).fillna("")
    # build maps
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()

    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)
    if not blocks:
        return
    global_dates = blocks[0][3]

    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        raw[0] = f"{surname}{two}" if matched else raw[0]
        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        # place affiliation in 31st column
        hdr[30] = rec_aff
        # extract digits from 8th column and set PH prefix in 30th column
        col8 = emp_col8_map.get(code, "")
        m = re.search(r"(\d+)", col8)
        if m:
            hdr[29] = f"PH{m.group(1)}"
        # header substitutions
        hdr = [
            re.sub(
                r"電話番号",
                "電話",
                re.sub(
                    r"社員番号",
                    "職番",
                    re.sub(r"PE([0-9]{6})", r"\1", re.sub(r"PE有効期限", "PE", v)),
                ),
            )
            for v in hdr
        ]
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            fe.append(
                [
                    clean_cell(df.iat[r2, j])
                    for r2 in range(d + 1, end)
                    if clean_cell(df.iat[r2, j])
                ]
            )
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))
        records.append(
            {
                "emp_no": code,
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )

    # pad entries
    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))

    # onboard
    for i, rec in enumerate(records):
        onb = []
        for idx, entries in enumerate(rec["full_entries"]):
            flights = [e for e in entries if re.match(r"^[0-9]", e)]
            names = []
            for j, other in enumerate(records):
                if i == j:
                    continue
                if any(f in other["full_entries"][idx] for f in flights):
                    names.append(other["hdr"][0])
            u = []
            for n in names:
                if n not in u:
                    u.append(n)
            onb.append(u)
        rec["onb"] = onb

    # remove duplicates
    seen = set()
    uniq = []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = uniq

    # sort by emp_order
    records.sort(
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        )
    )

    # output CSV
    out_csv = "formatted_schedule.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec["onb"]])
    # output Excel
    out_xlsx = "formatted_schedule19.xlsx"
    write_to_excel(records, emp_aff_map, out_xlsx)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--emp", default="emp_no.csv")
    a = p.parse_args()
    run(a.schedule, a.emp)
