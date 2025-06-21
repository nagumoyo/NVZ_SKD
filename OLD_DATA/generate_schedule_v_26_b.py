import streamlit as st
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill

st.title("整形済スケジュール V26d（空行/OB行除去・同乗者再計算）")
st.markdown("### ファイルをアップロードしてください")

schedule_file = st.file_uploader("スケジュールCSVファイル", type="csv")
emp_file = st.file_uploader("職員情報ファイル (emp_no.csv)", type="csv")

if st.button("出力"):
    if not schedule_file or not emp_file:
        st.error("両方のファイルをアップロードしてください。")
        st.stop()

    # 職員情報読み込み
    emp_df = pd.read_csv(emp_file, header=None, encoding="utf-8", keep_default_na=False)
    emp_df.columns = [f"col_{i+1}" for i in range(emp_df.shape[1])]
    emp_df["emp_no"] = emp_df["col_3"].astype(str).str.zfill(5)
    emp_df["full_name"] = emp_df["col_5"].str.strip() + emp_df["col_7"].str.strip()
    emp_df = emp_df.reset_index().rename(columns={"index": "sort_index"})

    # スケジュール読み込み
    schedule_df = pd.read_csv(schedule_file, header=None, dtype=str, encoding="utf-8")
    schedule_df.fillna("", inplace=True)

    # 前処理：空行とOB行削除
    def is_blank(row):
        return all(str(c).strip() == "" for c in row)

    def is_OB(row):
        first = str(row.iloc[0]).strip()
        if re.match(r"^[A-Z]+OB$", first):
            return True
        if re.search(r"00099\d{3}", "".join(row.astype(str))):
            return True
        return False

    mask = [not (is_blank(row) or is_OB(row)) for _, row in schedule_df.iterrows()]
    schedule_df = schedule_df.loc[mask].reset_index(drop=True)

    # 年月取得
    ym = schedule_df.iat[0, 4].strip()
    year, month = int(ym[:4]), int(ym[4:6])

    # ヘッダー行検出
    def is_header(i):
        cell = str(schedule_df.iat[i, 0]).strip()
        if not re.match(r"^[A-Z]{2,}$", cell):
            return False
        nxt = schedule_df.iloc[i + 1]
        cnt = sum(
            1 for c in nxt if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(c).strip())
        )
        return cnt >= 25

    header_idxs = [i for i in range(len(schedule_df) - 1) if is_header(i)]

    # クルーブロック
    blocks = []
    for idx, h in enumerate(header_idxs):
        nxt = header_idxs[idx + 1] if idx + 1 < len(header_idxs) else len(schedule_df)
        blocks.append((h, h + 1, nxt))

    # 各クルー情報
    all_crew = []
    for h, d, e in blocks:
        hdr_row = schedule_df.iloc[h]
        emp_no = None
        for c in hdr_row:
            m = re.search(r"000(\d{5})", str(c))
            if m:
                emp_no = m.group(1)
                break
        name = str(hdr_row.iat[0]).strip()
        phase = depart = ""
        sort_idx = 999999
        if emp_no:
            mr = emp_df[emp_df.emp_no == emp_no]
            if not mr.empty:
                name = mr.iat[0, emp_df.columns.get_loc("full_name")]
                phase = (
                    mr.iat[0, emp_df.columns.get_loc("col_8")]
                    if "col_8" in emp_df
                    else ""
                )
                depart = mr.iat[0, emp_df.columns.get_loc("col_1")]
                sort_idx = int(mr.iat[0, emp_df.columns.get_loc("sort_index")])
        # header list
        hdr = [name, "", f"000{emp_no}" if emp_no else ""] + [
            c for i, c in enumerate(hdr_row) if i not in (0, 2) and str(c).strip()
        ]
        hdr = hdr[:31] + [""] * (31 - len(hdr))
        hdr[29] = f"PH{phase}" if phase else ""
        hdr[30] = depart
        # dates and date_cols
        date_row = schedule_df.iloc[d]
        date_cols = [
            i
            for i, c in enumerate(date_row)
            if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(c).strip())
        ][:31]
        dr = [date_row.iat[i] for i in date_cols] + [""] * (31 - len(date_cols))
        # flight text
        texts = schedule_df.iloc[d + 1 : e]
        flight = texts.apply(
            lambda row: [
                (
                    str(row.iat[i])
                    if str(row.iat[i]).strip()
                    and re.match(r"^\d", str(row.iat[i]).strip())
                    else ""
                )
                for i in date_cols
            ],
            axis=1,
        )
        sched_m = ["\n".join(x for x in col if x) for col in zip(*flight.tolist())]
        all_crew.append({"hdr": hdr, "dr": dr, "sched": sched_m, "sort": sort_idx})

    # ソート
    all_crew = sorted(all_crew, key=lambda x: x["sort"])
    # 同乗再計算
    for crew in all_crew:
        onb = [""] * 31
        for idx_col in range(31):
            val = crew["sched"][idx_col]
            if re.match(r"^\d", val):
                co = [
                    c["hdr"][0]
                    for c in all_crew
                    if c is not crew and c["sched"][idx_col] == val
                ]
                onb[idx_col] = "\n".join(co)
        crew["onb"] = onb

    # Excel出力
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"
    thin = Side(border_style="thin", color="FF000000")
    br = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c in all_crew:
        ws.append(c["hdr"])
        [
            setattr(cell, "border", br)
            or setattr(
                cell, "alignment", Alignment(horizontal="center", vertical="center")
            )
            for cell in ws[ws.max_row]
        ]
        ws.append(c["dr"])
        [
            setattr(cell, "border", br)
            or setattr(
                cell, "alignment", Alignment(horizontal="center", vertical="center")
            )
            or setattr(
                cell, "fill", PatternFill(fill_type="solid", start_color="FFFFCC")
            )
            for cell in ws[ws.max_row]
        ]
        ws.append(c["sched"])
        [
            setattr(cell, "border", br)
            or setattr(
                cell,
                "alignment",
                Alignment(wrap_text=True, horizontal="center", vertical="top"),
            )
            for cell in ws[ws.max_row]
        ]
        ws.append(c["onb"])
        [
            setattr(cell, "border", br)
            or setattr(
                cell,
                "alignment",
                Alignment(wrap_text=True, horizontal="center", vertical="top"),
            )
            for cell in ws[ws.max_row]
        ]
    buf = io.BytesIO()
    wb.save(buf)
    st.download_button(
        "Download",
        buf.getvalue(),
        file_name=f"schedule_{year}_{month}_v26d.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
