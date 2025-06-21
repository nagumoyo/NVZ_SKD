import streamlit as st
import pandas as pd
import re
import io
import os
import string
import datetime
from openpyxl.styles import Border, Side, Alignment, PatternFill

st.title("整形済スケジュール V26a（日付行のみ色・空行/OB行削除）")

st.markdown("### ファイルをアップロードしてください")

schedule_file = st.file_uploader("スケジュールCSVファイル", type="csv")
emp_file = st.file_uploader("職員情報ファイル (emp_no.csv)", type="csv")

if st.button("出力"):
    if schedule_file is None or emp_file is None:
        st.error("スケジュールファイルと職員情報ファイルの両方をアップロードしてください。")
    else:
        # --- 職員情報読み込み ---
        emp_df = pd.read_csv(emp_file, header=None, encoding="utf-8", keep_default_na=False)
        emp_df.columns = [f"col_{i+1}" for i in range(emp_df.shape[1])]
        emp_df["emp_no"] = emp_df["col_3"].astype(str).str.zfill(5)
        emp_df["full_name"] = emp_df["col_5"].fillna("").str.strip() + emp_df["col_7"].fillna("").str.strip()
        emp_df = emp_df.reset_index().rename(columns={"index": "sort_index"})

        # --- スケジュール読み込み ---
        schedule_df = pd.read_csv(schedule_file, header=None, encoding="utf-8", dtype=str)
        schedule_df.fillna("", inplace=True)

        # ↓ ここから v26a 追加: 空行と OB 行の削除 ↓
        # 1) 全セルが空白文字のみの行を削除
        blank_rows = schedule_df.applymap(lambda x: str(x).strip() == "").all(axis=1)
        schedule_df = schedule_df.loc[~blank_rows].reset_index(drop=True)

        # 2) OB行を削除 (先頭セルが大文字+OB で終わる or 行内に00099xxxが含まれる)
        import re as _re
        def is_OB_initial(row):
            first = str(row.iloc[0]).strip()
            if _re.match(r"^[A-Z]+OB$", first):
                return True
            if _re.search(r"00099\d{3}", "".join(row.astype(str))):
                return True
            return False
        ob_mask = schedule_df.apply(is_OB_initial, axis=1)
        schedule_df = schedule_df.loc[~ob_mask].reset_index(drop=True)
        # ↑ ここまで追加 ↑

        # --- 元のスクリプトロジック（ヘッダー検出～出力） ---
        year_month_str = schedule_df.iloc[0, 4].strip()
        year = int(year_month_str[:4])
        month = int(year_month_str[4:6])

        def is_OB_row(row):
            first_cell = row[0].strip()
            if re.match(r"^[A-Z]+OB$", first_cell):
                return True
            for cell in row:
                if re.search(r"(00099\d{3})", str(cell)):
                    return True
            return False

        # 有効なヘッダー行インデックス収集
        header_indices_valid = []
        for i in range(schedule_df.shape[0] - 1):
            first_cell = schedule_df.iloc[i, 0].strip()
            next_row = schedule_df.iloc[i + 1, :].tolist()
            header_match = bool(re.match(r"^[A-Z]{2,}$|^[A-Z]{3,}$", first_cell))
            date_count = sum(1 for cell in next_row if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip()))
            if header_match and date_count >= 25:
                header_indices_valid.append(i)

        # クルーブロック切り出し
        crew_blocks = []
        for idx, header_idx in enumerate(header_indices_valid):
            next_header_idx = header_indices_valid[idx + 1] if idx + 1 < len(header_indices_valid) else schedule_df.shape[0]

            block_end = schedule_df.shape[0]
            for j in range(header_idx + 2, schedule_df.shape[0]):
                row = schedule_df.iloc[j, :].tolist()
                first_cell = row[0].strip()
                next_row = schedule_df.iloc[j + 1, :].tolist() if j + 1 < schedule_df.shape[0] else []

                is_fake_header = bool(re.match(r"^[A-Z]{2,}$|^[A-Z]{3,}$", first_cell)) and \
                                 (sum(1 for cell in next_row if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip())) < 25)

                is_blank_row = all([str(cell).strip() == "" for cell in row])
                is_OB_or_end = is_OB_row(row) or is_fake_header or is_blank_row

                if is_OB_or_end or j >= next_header_idx - 1:
                    block_end = max(j - 1, header_idx + 2)
                    break

            header_row = schedule_df.iloc[header_idx, :].tolist()
            if is_OB_row(header_row):
                continue
            crew_blocks.append((header_idx, header_idx + 1, block_end))

        # 以下、全クルーのスケジュールマージ・ソート・出力処理（元 v25 のまま）
        output_rows = []
        all_crew_schedules = []
        row_counter = 1

        for block in crew_blocks:
            header_raw = schedule_df.iloc[block[0], :].tolist()
            emp_no_match = next(
                (re.search(r"(000\d{5})", cell).group(1)[3:]
                 for cell in header_raw if re.search(r"(000\d{5})", cell)),
                None
            )

            emp_name = ""
            fo_phase_value = ""
            depart_value = ""
            sort_index = 999999

            if emp_no_match:
                match_row = emp_df.loc[emp_df["emp_no"] == emp_no_match]
                if not match_row.empty:
                    emp_name = match_row.iloc[0]["full_name"]
                    fo_phase_value = match_row.iloc[0]["col_8"] if "col_8" in match_row.columns else ""
                    depart_value = match_row.iloc[0]["col_1"]
                    sort_index = int(match_row.iloc[0]["sort_index"])

            name_first_cell = header_raw[0].strip()
            final_name = emp_name if emp_name else name_first_cell

            header_elements = [final_name, "", "000" + emp_no_match if emp_no_match else ""]
            header_elements += [cell for i, cell in enumerate(header_raw) if i not in [0, 2] and str(cell).strip() != ""]

            date_row = schedule_df.iloc[block[1], :].tolist()

            sched_search_range = schedule_df.iloc[block[1] + 1: block[2] + 1, :]
            block_start_dynamic_idx = None
            for offset, (_, row) in enumerate(sched_search_range.iterrows()):
                if not all([str(cell).strip() == "" for cell in row.tolist()]):
                    block_start_dynamic_idx = block[1] + 1 + offset
                    break
            if block_start_dynamic_idx is not None:
                sched_rows = schedule_df.iloc[block_start_dynamic_idx: block[2] + 1, :]
            else:
                sched_rows = pd.DataFrame()

            date_col_indices = [i for i, cell in enumerate(date_row) if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip())]
            date_col_indices = date_col_indices[:31]

            header_row_fixed = [header_elements[i] if i < len(header_elements) else "" for i in range(31)]
            if len(header_row_fixed) < 31:
                header_row_fixed += [""] * (31 - len(header_row_fixed))
            header_row_fixed[29] = "PH" + fo_phase_value if fo_phase_value else ""
            header_row_fixed[30] = depart_value if depart_value else ""

            date_row_fixed = [date_row[i] for i in date_col_indices]
            while len(date_row_fixed) < 31:
                date_row_fixed.append("")

            merged_schedule_row = [""] * 31
            for i, col_idx in enumerate(date_col_indices):
                sched_texts = sched_rows.iloc[:, col_idx].apply(lambda x: str(x)).tolist() if not sched_rows.empty else []
                merged_schedule_row[i] = "\n".join(sched_texts)

            all_crew_schedules.append({
                'name': final_name,
                'header_row': header_row_fixed,
                'date_row': date_row_fixed,
                'schedule_row': merged_schedule_row,
                'row_number': row_counter + 1,
                'sort_index': sort_index
            })
            row_counter += 4

        all_crew_schedules = sorted(all_crew_schedules, key=lambda x: x['sort_index'])

        # Streamlit で Excel 出力
        output = io.BytesIO()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月"

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for crew in all_crew_schedules:
            # ヘッダー行
            ws.append(crew['header_row'])
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 日付行
            ws.append(crew['date_row'])
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type="solid", start_color="FFFFCC")
            # スケジュール行
            ws.append(crew['schedule_row'])
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")

        wb.save(output)
        st.download_button(
            label="整形済スケジュールをダウンロード",
            data=output.getvalue(),
            file_name=f"schedule_{year}_{month}_v26a.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
