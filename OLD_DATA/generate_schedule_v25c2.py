import streamlit as st
import pandas as pd
import re
import io
import os
import string
import datetime
from openpyxl.styles import Border, Side, Alignment, PatternFill

st.title("整形済スケジュール V25（日付行のみ色）")

st.markdown("### ファイルをアップロードしてください")

schedule_file = st.file_uploader("スケジュールCSVファイル", type="csv")
emp_file = st.file_uploader("職員情報ファイル (emp_no.csv)", type="csv")

if st.button("出力"):
    if schedule_file is None or emp_file is None:
        st.error("スケジュールファイルと職員情報ファイルの両方をアップロードしてください。")
    else:
        emp_df = pd.read_csv(emp_file, header=None, encoding="utf-8", keep_default_na=False)
        emp_df.columns = [f"col_{i+1}" for i in range(emp_df.shape[1])]
        emp_df["emp_no"] = emp_df["col_3"].astype(str).str.zfill(5)
        emp_df["full_name"] = emp_df["col_5"].fillna("").str.strip() + emp_df["col_7"].fillna("").str.strip()

        emp_df = emp_df.reset_index().rename(columns={"index": "sort_index"})

        schedule_df = pd.read_csv(schedule_file, header=None, encoding="utf-8", dtype=str)
        schedule_df.fillna("", inplace=True)

        year_month_str = schedule_df.iloc[0, 4].strip()
        year = int(year_month_str[:4])
        month = int(year_month_str[4:6])

def is_OB_row(row):
    """ OB行（退職・異動などのOB行）判定ルーチン """
            first_cell = row[0].strip()
            if re.match(r"^[A-Z]+OB$", first_cell):
                return True
            for cell in row:
                if re.search(r"(00099\d{3})", str(cell)):
                    return True
            return False

        header_indices_valid = []
        for i in range(schedule_df.shape[0] - 1):
            first_cell = schedule_df.iloc[i, 0].strip()
            next_row = schedule_df.iloc[i + 1, :].tolist()
            header_match = bool(re.match(r"^[A-Z]{2,}$|^[A-Z]{3,}$", first_cell))
            date_count = sum(1 for cell in next_row if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip()))
            if header_match and date_count >= 25:
                header_indices_valid.append(i)

        crew_blocks = []
        for idx, header_idx in enumerate(header_indices_valid):
            next_header_idx = header_indices_valid[idx + 1] if idx + 1 < len(header_indices_valid) else schedule_df.shape[0]

            block_end = schedule_df.shape[0]
            for j in range(header_idx + 2, schedule_df.shape[0]):
                row = schedule_df.iloc[j, :].tolist()
                first_cell = row[0].strip()
                next_row = schedule_df.iloc[j + 1, :].tolist() if j + 1 < schedule_df.shape[0] else []

                is_fake_header = bool(re.match(r"^[A-Z]{2,}$|^[A-Z]{3,}$", first_cell)) and \
                                 (sum(1 for cell in next_row if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip())) < 25)

                is_blank_row = all([str(cell).strip() == "" for cell in row])
                is_OB_or_end = is_OB_row(row) or is_fake_header or is_blank_row

                if is_OB_or_end or j >= next_header_idx - 1:
                    block_end = max(j - 1, header_idx + 2)
                    break

            header_row = schedule_df.iloc[header_idx, :].tolist()
            if is_OB_row(header_row):
                continue
            # crew_blocks に [header行 index, 日付行 index, block_end行 index] を追加
            crew_blocks.append((header_idx, header_idx + 1, block_end))
        output_rows = []
        all_crew_schedules = []
        row_counter = 1

        for block in crew_blocks:
            header_raw = schedule_df.iloc[block[0], :].tolist()
            emp_no_match = next(
                (re.search(r"(000\d{5})", cell).group(1)[3:]
                 for cell in header_raw if re.search(r"(000\d{5})", cell)),
                None
            )

            emp_name = ""
            fo_phase_value = ""
            depart_value = ""
            sort_index = 999999

            if emp_no_match:
                match_row = emp_df.loc[emp_df["emp_no"] == emp_no_match]
                if not match_row.empty:
                    emp_name = match_row.iloc[0]["full_name"]
                    fo_phase_value = match_row.iloc[0]["col_8"]
                    depart_value = match_row.iloc[0]["col_1"]
                    sort_index = int(match_row.iloc[0]["sort_index"])

            name_first_cell = header_raw[0].strip()
            final_name = emp_name if emp_name else name_first_cell

            header_elements = [final_name, "", "000" + emp_no_match if emp_no_match else ""]
            header_elements += [cell for i, cell in enumerate(header_raw) if i not in [0, 2] and str(cell).strip() != ""]

            date_row = schedule_df.iloc[block[1], :].tolist()

            sched_search_range = schedule_df.iloc[block[1] + 1: block[2] + 1, :]
            block_start_dynamic_idx = None
            for offset, (_, row) in enumerate(sched_search_range.iterrows()):
                if not all([str(cell).strip() == "" for cell in row.tolist()]):
                    block_start_dynamic_idx = block[1] + 1 + offset
                    break
            if block_start_dynamic_idx is not None:
                sched_rows = schedule_df.iloc[block_start_dynamic_idx: block[2] + 1, :]
            else:
                sched_rows = pd.DataFrame()

            date_col_indices = [i for i, cell in enumerate(date_row) if re.match(r"^(0?[1-9]|[12][0-9]|3[01])$", str(cell).strip())]
            date_col_indices = date_col_indices[:31]

            header_row_fixed = [header_elements[i] if i < len(header_elements) else "" for i in range(31)]

            if len(header_row_fixed) < 31:
                header_row_fixed += [""] * (31 - len(header_row_fixed))
            header_row_fixed[29] = "PH" + fo_phase_value if fo_phase_value else ""
            header_row_fixed[30] = depart_value if depart_value else ""

            date_row_fixed = [date_row[i] for i in date_col_indices]
            while len(date_row_fixed) < 31:
                date_row_fixed.append("")

            merged_schedule_row = [""] * 31
            for i, col_idx in enumerate(date_col_indices):
                sched_texts = sched_rows.iloc[:, col_idx].apply(lambda x: str(x)).tolist() if not sched_rows.empty else []
                merged_schedule_row[i] = "\n".join(sched_texts)

            all_crew_schedules.append({
                'name': final_name,
                'header_row': header_row_fixed,
                'date_row': date_row_fixed,
                'schedule_row': merged_schedule_row,
                'row_number': row_counter + 1,
                'sort_index': sort_index
            })

            row_counter += 4

        all_crew_schedules = sorted(all_crew_schedules, key=lambda x: x['sort_index'])

        for crew in all_crew_schedules:
            my_name = crew['name']
            header_row_fixed = crew['header_row']
            date_row_fixed = crew['date_row']
            my_sched_row = crew['schedule_row']
            my_row_number = crew['row_number']
            onboard_row = [""] * 31

            for i in range(len(my_sched_row)):
                my_sched_parts = my_sched_row[i].strip().split("\n")
                same_sched_crew = set()

                for my_sched_part in my_sched_parts:
                    my_prefix = re.match(r"^(\d+)", my_sched_part.strip())
                    if my_prefix:
                        my_prefix_num = my_prefix.group(1)

                        for other_crew in all_crew_schedules:
                            if other_crew['name'] == my_name:
                                continue
                            if i < len(other_crew['schedule_row']):
                                other_sched_parts = other_crew['schedule_row'][i].strip().split("\n")
                                for other_sched_part in other_sched_parts:
                                    other_prefix = re.match(r"^(\d+)", other_sched_part.strip())
                                    if other_prefix and other_prefix.group(1) == my_prefix_num:
                                        same_sched_crew.add(other_crew['name'])

                if same_sched_crew:
                    onboard_row[i] = "\n".join(sorted(same_sched_crew))
                else:
                    onboard_row[i] = ""

            output_rows.append(header_row_fixed)
            output_rows.append(date_row_fixed)
            output_rows.append(my_sched_row)
            output_rows.append(onboard_row)

        def clean_column_name(cell):
            cell = cell.replace("社員番号", "社番")
            cell = cell.replace("電話番号", "電話")
            cell = cell.replace("PE有効期限", "PE")
            cell = re.sub(r"PE([0-9]{6})", r"\1", cell)
            return cell

        for i in range(0, len(output_rows), 4):
            output_rows[i] = [clean_column_name(str(cell)) for cell in output_rows[i]]

        output_df = pd.DataFrame(output_rows)

        base_name = "output_schedule25"
        ext = ".xlsx"

        for suffix in string.ascii_lowercase:
            output_file = f"{base_name}{suffix}{ext}"
            if not os.path.exists(output_file):
                break

        double_border = Border(
            top=Side(style='double'),
            bottom=Side(style='double')
        )

        alignment_wrap_top = Alignment(
            vertical='top',
            wrap_text=True
        )

        alignment_no_wrap = Alignment(
            vertical='top',
            wrap_text=False
        )

        phone_pattern = re.compile(r"0[0-9]{2,}-[0-9]{3,}-[0-9]{4}")

        saturday_fill = PatternFill(start_color="C1E4E9", end_color="C1E4E9", fill_type="solid")
        sunday_fill = PatternFill(start_color="F6B894", end_color="F6B894", fill_type="solid")

        column_weekdays = []
        for day_str in all_crew_schedules[0]["date_row"]:
            if day_str.strip().isdigit():
                day = int(day_str)
                try:
                    weekday = datetime.date(year, month, day).weekday()
                    column_weekdays.append(weekday)
                except:
                    column_weekdays.append(None)
            else:
                column_weekdays.append(None)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, header=False)
            worksheet = writer.sheets['Sheet1']

            for row_idx in range(1, output_df.shape[0] + 1):
                for col_idx in range(1, 32):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = str(cell.value) if cell.value else ""

                    if (row_idx - 1) % 4 == 0:
                        cell.border = double_border

                    if phone_pattern.fullmatch(cell_value):
                        cell.alignment = alignment_no_wrap
                    else:
                        cell.alignment = alignment_wrap_top

                    weekday = column_weekdays[col_idx - 1]

                    # ★ 日付行だけ色をつける！
                    if (row_idx - 1) % 4 == 1:
                        if weekday == 5:
                            cell.fill = saturday_fill
                        elif weekday == 6:
                            cell.fill = sunday_fill

        buffer.seek(0)

        st.success("出力が完了しました！ 下のボタンからダウンロードしてください。")
        st.download_button(
            label="📥 ダウンロード",
            data=buffer,
            file_name=output_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )