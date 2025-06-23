# === generate_schedule.py ===
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ====


def clean_cell(text):
    """Remove invisible characters and trim whitespace"""
    s = str(text)
    return re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s).strip()


def remove_blank_and_ob(df):
    """Drop rows that are all blank or contain only 'OB'"""
    df = df.replace({pd.NA: "", "nan": "", "NaN": ""})
    df = df[(df != "").any(axis=1)]
    df = df[~df.apply(lambda row: all(str(c).strip() == "OB" for c in row), axis=1)]
    return df.reset_index(drop=True)


def find_header_rows(df):
    """Detect header rows: first column ends with two uppercase letters and next row contains a date"""
    hdrs = []
    for i in range(len(df) - 1):
        if re.search(r"[A-Z]{2}$", clean_cell(df.iat[i, 0])):
            next_row = [clean_cell(x) for x in df.iloc[i + 1]]
            if any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row):
                hdrs.append(i)
    return hdrs


def slice_blocks(df):
    """Slice DataFrame into blocks defined by header rows"""
    hdrs = find_header_rows(df)
    blocks = []
    N = len(df)
    for idx, h in enumerate(hdrs):
        nxt = hdrs[idx + 1] if idx + 1 < len(hdrs) else N
        for i in range(h + 1, nxt):
            row = [clean_cell(x) for x in df.iloc[i]]
            dates = [
                j
                for j, v in enumerate(row)
                if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
            ]
            if dates:
                blocks.append((h, i, nxt, dates))
                break
    return blocks


def write_to_excel(records, out_xlsx):
    wb = Workbook()
    ws = wb.active
    grey = PatternFill(fill_type="solid", fgColor="DDDDDD")
    hi = PatternFill(fill_type="solid", fgColor="FFFF00")
    dbl = Side(border_style="double", color="000000")
    brd = Border(top=dbl, bottom=dbl)

    # Set column widths
    for c in range(1, 32):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 8

    r = 1
    for rec in records:
        # Header row
        for c, v in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.border = brd
        r += 1

        # Date row
        for c, v in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = grey
        r += 1

        # Schedule row
        for c, v in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        r += 1

        # Onboard row
        for c, v in enumerate(rec["onb"], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if "*" in v:
                cell.fill = hi
        r += 1

    wb.save(out_xlsx)


def run(schedule_input, emp_input, config_path=None):
    # Read input CSVs
    sched_df = pd.read_csv(schedule_input, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_input, header=None, dtype=str).fillna("")

    # Build employee map (emp_no -> name)
    emp_map = {row[2]: row[3] for _, row in emp_df.iterrows() if row[2] and row[3]}

    # Clean and slice into blocks
    df = sched_df.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)

    crew_data = []
    for h, d, e, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        hdr_vals = [v for v in raw if v]
        hdr = hdr_vals[:31] + [""] * (31 - len(hdr_vals[:31]))

        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))

        # All entries for schedule display
        full_entries = [
            [
                clean_cell(df.iat[r, j])
                for r in range(d + 1, e)
                if clean_cell(df.iat[r, j])
            ]
            for j in dates
        ]

        # Numeric-only entries for onboard matching
        flight_entries = [
            [entry for entry in events if re.match(r"^[0-9]", entry)]
            for events in full_entries
        ]

        emp_no = None
        if len(hdr) > 2:
            m = re.search(r"(\d{5})$", hdr[2])
            emp_no = m.group(1) if m else None

        crew_name = emp_map.get(emp_no, hdr_vals[0] if hdr_vals else "")

        crew_data.append(
            {
                "hdr": hdr,
                "dr": dr,
                "full_entries": full_entries,
                "flight_entries": flight_entries,
                "emp_no": emp_no,
                "crew_name": crew_name,
            }
        )

    # Determine onboard names
    for i, cd in enumerate(crew_data):
        onb = []
        for idx, flights in enumerate(cd["flight_entries"]):
            names = []
            for j, other in enumerate(crew_data):
                if i == j:
                    continue
                if idx < len(other["flight_entries"]):
                    of = other["flight_entries"][idx]
                    if any(f in of for f in flights):
                        names.append(other["crew_name"])
            onb.append("\n".join(names))
        onb += [""] * (31 - len(onb))
        cd["sched"] = ["\n".join(lst) for lst in cd["full_entries"]] + [""] * (
            31 - len(cd["full_entries"])
        )
        cd["onb"] = onb

    records = [
        {"hdr": cd["hdr"], "dr": cd["dr"], "sched": cd["sched"], "onb": cd["onb"]}
        for cd in crew_data
    ]

    # CSV output
    out_csv = "formatted_schedule.csv"
    rows = []
    for rec in records:
        rows.extend([rec["hdr"], rec["dr"]] + rec["sched"] + [rec["onb"]])
    pd.DataFrame(rows).to_csv(out_csv, index=False, header=False)

    # Excel output
    out_xlsx = "formatted_schedule.xlsx"
    write_to_excel(records, out_xlsx)

    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--schedule", default="schedule.csv")
    parser.add_argument("--emp", default="emp_no.csv")
    args = parser.parse_args()
    run(args.schedule, args.emp)
