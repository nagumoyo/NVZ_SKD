#!/usr/bin/env python3
# === generate_schedule31b.py ===

import pandas as pd
import re
import csv
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import logging

# # ロガー設定（冒頭に記載）
# logging.basicConfig(
#     filename="simslot_debug.log",
#     level=logging.DEBUG,
#     format="%(asctime)s [%(levelname)s] %(message)s",
# )
# logger = logging.getLogger(__name__)

# ------------------------------------------
# 警告抑制設定
# ------------------------------------------
# openpyxl のスタイルシート関連の警告をまとめて無視
warnings.filterwarnings(
    "ignore", category=UserWarning, module="openpyxl.styles.stylesheet"
)

# ── Helpers: 補助関数群 ──────────────────────────────────────────────────


def load_sim_slot_excel(file):
    """
    SIM Slot Excel ファイルを pandas.DataFrame として読み込むユーティリティ関数。

    本関数はローカル実行と Streamlit 実行の両方に対応:
      1. file が文字列の場合 → ローカルファイルパスを直接読み込む
      2. それ以外 (Streamlit UploadedFile) → BytesIO 経由で読み込む

    Args:
        file (str or UploadedFile):
            - str: ローカル環境での Excel ファイルパス
            - UploadedFile: Streamlit アップロードオブジェクト

    Returns:
        pandas.DataFrame:
            - 読み込みに成功した場合: SIM Slot のデータ
            - 失敗した場合: 空の DataFrame
    """
    import pandas as _pd
    import io as _io

    try:
        if isinstance(file, str):
            # ローカルパスから直接読み込み (ヘッダーは3行目)
            df = _pd.read_excel(file, header=2)
        else:
            # StreamlitのアップロードオブジェクトをBytesIO経由で読み込み
            raw = file.read()
            df = _pd.read_excel(_io.BytesIO(raw), header=2)

        # 読み込んだ行数をログとして表示
        print(f"[INFO] SIM Slot: {len(df)} 行を読み込みました")
        return df

    except Exception as e:
        # 読み込み失敗時はエラーログを出力し、空 DataFrame を返却
        print(f"[ERROR] SIM Slot Excel 読み込み失敗: {e}")
        return _pd.DataFrame()


def load_simslot_schedule_codes(
    pref_file_path="PREF.xlsx", sheet_name="SIMSLOT"
) -> list:
    """
    PREF.xlsx の SIMSLOT シートから ActivityTypeCode のリストを取得する。

    - シート名が存在しない場合は空リストを返却
    - B1 セルにカンマ区切りでコードが設定されている想定

    Args:
        pref_file_path (str): 設定ファイルのパス
        sheet_name (str): 読み込むシート名 (デフォルト "SIMSLOT")

    Returns:
        list: 有効なコードのリスト
    """
    try:
        wb = load_workbook(pref_file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"✅ SIMSLOT シートが存在しません: {sheet_name}")
            return []

        ws = wb[sheet_name]
        b1_value = ws["B1"].value
        if not b1_value:
            print("✅ B1セルが空です")
            return []

        # カンマ区切りの文字列を分割し、前後空白を除去
        return [code.strip() for code in str(b1_value).split(",") if code.strip()]

    except Exception as e:
        print(f"⚠️ SIMSLOT 読み込み失敗: {e}")
        return []


def clean_cell(text: any) -> str:
    """
    セル内の文字列を正規化・クリーンアップする関数。

    処理内容:
      1. NaN / None を検出して空文字列に変換
      2. ゼロ幅スペースや NBSP、タブ、改行などの不可視文字を削除
      3. 文字列の前後の余分なスペースをトリム
      4. 最終的に空白のみの文字列は空文字列に統一

    Args:
        text: 任意の型で渡されるセルの元データ

    Returns:
        str: クリーンアップ後の文字列
    """
    import pandas as _pd
    import re as _re

    # 1) NaN / None を空文字に
    if _pd.isna(text):
        return ""
    s = str(text)
    # 2) 不可視文字や改行をすべて削除
    s = _re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s)
    # 3) 前後のスペースをトリム
    s = s.strip()
    # 4) 空文字か判定
    return s if s else ""


def remove_blank_and_ob(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame から不要な行を除去する関数。

    - 完全に空文字列の行を削除
    - 全セルが "" または "OB" のみの行を削除（Out-of-Bounds を想定）
    - 特定の Emp ID を含む行も除外可能（skip_ids を設定）

    Args:
        df: セルが文字列化され、クリーニング済みの pandas DataFrame

    Returns:
        pd.DataFrame: 空行と OB 行、特定 Emp ID 行を除去した DataFrame
    """
    # 除外する Emp ID のセット (例として3IDを設定)
    skip_ids = {"00049292", "00035957", "00041169"}

    rows = []
    for row in df.values:
        texts = [str(x).strip() for x in row]

        # 1) OB の既存ロジック: 99xxx 系や全 OB 行、末尾に OB が続く行を除去
        if any(re.fullmatch(r"00099[0-9]{3}", t) for t in texts):
            continue
        if all(not t or t == "OB" for t in texts):
            continue
        if re.fullmatch(r"[A-Z]+OB", texts[0]):
            continue

        # 2) 追加: 特定 Emp ID を含む行をスキップ
        if any(t in skip_ids for t in texts):
            continue

        rows.append(row)

    return pd.DataFrame(rows, columns=df.columns)


def find_header_rows(df: pd.DataFrame) -> list:
    """
    ヘッダー行を検出する関数。

    - 列0: 姓名+2レターが大文字英字で構成されるセル
    - 列2: 2レター (2文字の大文字英字)
    - 次行に日付行 (01～31の数値) が存在する

    Args:
        df (pd.DataFrame): クリーン済みの DataFrame

    Returns:
        list: ヘッダー行のインデックス一覧
    """
    hdrs = []
    for i in range(len(df) - 1):
        # 現行の1列目と3列目（0-indexed）をチェック
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        # 次行をリスト化して日付候補を確認
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]

        # 条件: c0 が 2文字以上大文字、c2 が 2文字大文字、次行に日付あり
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    return hdrs


def slice_blocks(df: pd.DataFrame) -> list:
    """
    ヘッダー／日付／スケジュール行ブロックを切り出す関数。

    ヘッダ行から次のヘッダ行までを1ブロックとして扱い、
    日付行位置とスケジュール列番号を算出する。

    Args:
        df (pd.DataFrame): クリーン済みの DataFrame

    Returns:
        list of tuples: (header_idx, date_idx, end_idx, date_columns)
    """
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)

    for idx, h in enumerate(hdrs):
        # 次のヘッダー行または末尾までを範囲とする
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        date_idx = h + 1
        # 最初の候補日付行が正しくなければ、次の行も探索
        vals = [str(df.iat[date_idx, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    date_idx = r
                    break
        # 日付列インデックス群を取得
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[date_idx, k]).strip() for k in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]

        blocks.append((h, date_idx, end, date_cols))
    return blocks


def load_pref_rules(file: str = "PREF.xlsx", sheet_name: str = "色分け設定") -> list:
    """
    色分けルールを Excel から読み込む。

    - シート2行目以降をルールとして取得
    - ENABLE が YES の行のみ有効
    - 6列目のセル背景色を取得し、カラーコードとして保存

    Args:
        file (str): 設定ファイルパス
        sheet_name (str): シート名

    Returns:
        list of dict: 各ルール辞書を返却
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] シート「{sheet_name}」が{file}に存在しません。")
        return []

    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=2, values_only=True))
    rules = []

    for idx, row in enumerate(data, start=2):
        enable, first, op, second, label, _ = row[:6]
        if str(enable).strip().upper() != "YES":
            continue
        # 6列目の背景色を取得
        cell = ws.cell(row=idx, column=6)
        color = (
            cell.fill.start_color.rgb
            if cell.fill
            and cell.fill.start_color
            and cell.fill.start_color.type == "rgb"
            else None
        )
        rules.append(
            {
                "first": str(first or "").strip(),
                "second": str(second or "").strip(),
                "op": str(op or "").strip().upper(),
                "label": str(label or "").strip(),
                "color": color or "",
            }
        )
    return rules


# ── 5) ルール適用ロジック ─────────────────────────────────────────────


def match_rule_in_multiline(
    text: any, rule: dict, debug_log_path: str = "debug_log.txt"
) -> bool:
    """
    複数行テキストに対し、PREFルールの条件1/条件2をAND/OR論理でチェックする。
    デバッグ情報はファイルに追記。

    Args:
        text (any): セル内の元テキスト（複数行想定）
        rule (dict): PREFから取得したルール辞書:
            {
              'first': str,  # 条件1
              'second': str, # 条件2
              'op': 'AND'|'OR'|'NONE'
            }
        debug_log_path (str): デバッグログ出力先パス
    Returns:
        bool: ルールにマッチすれば True
    """
    lines = str(text).split("\n")
    cond1 = rule.get("first", "")
    cond2 = rule.get("second", "")
    op = rule.get("op", "NONE").upper()

    cond1_found = any(cond1 in line for line in lines) if cond1 else False
    cond2_found = any(cond2 in line for line in lines) if cond2 else False

    # # デバッグ情報をログファイルに追記
    # with open(debug_log_path, "a", encoding="utf-8") as f:
    #     f.write("\n=== DEBUG MATCH ===\n")
    #     f.write(f"Text Lines: {lines}\n")
    #     f.write(f"Rule: {rule}\n")
    #     f.write(f"cond1_found={cond1_found}, cond2_found={cond2_found}, op={op}\n")

    if op == "AND":
        return cond1_found and cond2_found
    if op == "OR":
        return cond1_found or cond2_found
    # NONE なら条件いずれかでマッチ
    return cond1_found or cond2_found


def apply_pref_rules_to_cell(
    cell, val: any, rules: list, fallback_color: str = None
) -> None:
    """
    セルの文字列に対し、PREFルールを適用して背景色を設定する。

    Args:
        cell (openpyxl.cell): Excel セルオブジェクト
        val (any): セルに表示するスケジュール文字列
        rules (list): load_pref_rules で取得したルールリスト
        fallback_color (str): マッチしなかった場合に適用するデフォルトカラー
    """
    text = str(val).strip()
    lines = [line.strip() for line in text.split("\n") if line.strip()]

    for idx, rule in enumerate(rules):
        if not isinstance(rule, dict):
            print(f"⚠️ 無効なルール形式 (index={idx}): {rule}")
            continue

        first = rule.get("first", "")
        second = rule.get("second", "")
        op = rule.get("op", "NONE").upper()

        # 行ごとに正規表現マッチ
        cond1 = (
            any(re.search(first, line) for line in lines)
            if first not in (None, "")
            else False
        )
        cond2 = (
            any(re.search(second, line) for line in lines)
            if second not in (None, "")
            else False
        )

        if (
            (op == "AND" and cond1 and cond2)
            or (op == "OR" and (cond1 or cond2))
            or (op == "NONE" and (cond1 or cond2))
        ):
            color = rule.get("color", "").lstrip("#")
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            return

    # どのルールにもマッチしない場合、フォールバック色を設定
    if fallback_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fallback_color)


# ── 6) Excel 出力: 行ごとの書込み ─────────────────────────────────────────


from openpyxl.styles import Alignment, PatternFill


def write_onboard_rows(
    ws,
    start_row: int,
    onboard_data: list[list[str]],
    emp_aff_map: dict,
    name_to_emp: dict,
    block_aff: str,
    self_name: str,
    name_to_row: dict,
    offset: int,
) -> int:
    """
    同乗クルー情報を Excel に行単位で書き込む（可変長＋ハイパーリンク＋オフセット対応）。

    Args:
        ws: openpyxl ワークシートオブジェクト
        start_row: 書込み開始行番号
        onboard_data: 各日付の同乗者名リストのリスト
        emp_aff_map: EmpNo→所属マップ
        name_to_emp: レコード名→EmpNo マップ
        block_aff: 現在のブロックの所属
        self_name: 本人の表示名
        name_to_row: 名前→Excel行番号マップ
        offset: 本体データを何列右にずらすか（例：5）
    Returns:
        int: 書き込んだ同乗行数
    """
    # 同乗者が最も多い日付に合わせた行数
    max_onb = max((len(day) for day in onboard_data if day), default=0)

    for i in range(max_onb):
        for j, names in enumerate(onboard_data, start=1):
            # i 行目の同乗者（なければ空文字）
            display = names[i] if i < len(names) else ""
            # 実際に書き込む列
            col_idx = j + offset
            # セルを取得
            cell = ws.cell(row=start_row + i, column=col_idx)

            # ハイパーリンク設定 or そのまま文字列
            if display:
                # 先頭の I/T プレフィックスを除いた raw 名
                raw = display[1:] if display.startswith(("I", "T")) else display
                target = name_to_row.get(raw)
                if target:
                    # シート名が "Schedule" の場合
                    cell.value = f'=HYPERLINK("#Schedule!A{target}", "{display}")'
                else:
                    cell.value = display
            else:
                cell.value = ""

            # 書式設定：左上揃え＋折り返し
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            # 同ブロック所属なら背景色ハイライト
            if display:
                emp_no = name_to_emp.get(raw)
                if emp_no and emp_aff_map.get(emp_no) == block_aff:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFEE99")

    return max_onb


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import re

# これをファイルの先頭あたりでimportしておいてください
# from your_module import apply_pref_rules_to_cell


def write_to_excel(
    records: list, emp_aff_map: dict, out_xlsx: str, pref_rules: list
) -> None:
    """
    全レコードを Excel ファイルに書き出す。

    - 先頭に5列追加 → A列: emp_no-行番号, B列: 英字ヘッダー名 + 2レター
    - ヘッダー行／日付行／スケジュール行／同乗行を順次配置
    - ヘッダー行に罫線、日付行に背景色、スケジュール行に折り返しなど
    """
    wb = Workbook()
    ws = wb.active

    # ① 先頭に5列挿入
    ws.insert_cols(1, amount=5)
    offset = 5

    # 同乗者リンク用マップ準備
    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}
    name_to_row = {}
    r = 1
    for rec in records:
        onb_cnt = sum(len(day) for day in rec.get("onb", []))
        height = 3 + onb_cnt
        r += height
        name_to_row[rec["hdr"][0]] = r

    current_row = 1
    for rec in records:
        # ── A/B列埋め ──
        emp_id = rec["emp_no"]
        hdr_alpha = rec["hdr"][0]
        raw_name = rec.get("orig_name", hdr_alpha[:-2])
        two = hdr_alpha[-2:]
        full_name = f"{raw_name} {two}"
        onb_cnt = sum(len(day) for day in rec.get("onb", []))
        block_height = 3 + onb_cnt

        for i in range(block_height):
            ws.cell(row=current_row + i, column=1, value=f"{emp_id}-{i+1}")
            ws.cell(row=current_row + i, column=2, value=full_name)
        # ── ヘッダー行 ──
        for col_idx, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(
                row=current_row,
                column=col_idx + offset,
                value=val,
            )
            # 罫線（二重線上・下）
            cell.border = Border(
                top=Side(border_style="double"),
                bottom=Side(border_style="double"),
            )
            # 折り返し or 縮小
            is_phone = bool(re.search(r"\d{2,4}-\d{2,4}-\d{4}", val)) or "電話" in val
            is_pe = bool(re.fullmatch(r"PE\d{6}", val))
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=not (is_phone or is_pe),
                shrink_to_fit=is_pe,
            )

        # ── 日付行 ──
        for col_idx, dv in enumerate(rec["dr"], start=1):
            cell = ws.cell(
                row=current_row + 1,
                column=col_idx + offset,
                value=dv,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            # 色分けルール適用
            apply_pref_rules_to_cell(
                cell,
                rec["sched"][col_idx - 1],
                pref_rules,
                fallback_color="DDDDDD",
            )

        # ── スケジュール行 ──
        for col_idx, sv in enumerate(rec["sched"], start=1):
            cell = ws.cell(
                row=current_row + 2,
                column=col_idx + offset,
                value=sv,
            )
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

        # ── 同乗者行 ──
        onboard_count = write_onboard_rows(
            ws,
            current_row + 3,
            rec.get("onb", []),
            emp_aff_map,
            name_to_emp,
            rec.get("aff"),
            rec["hdr"][0],
            name_to_row,
            offset,
        )

        current_row += 3 + onboard_count

    # 保存
    wb.save(out_xlsx)


import pandas as pd
import re
import csv
import os
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def run(
    schedule_file, pref_file="PREF.xlsx", sim_file="SIM Slot List 202507.xlsx"
) -> tuple[str, str]:
    # スケジュール CSV 読み込み（str or UploadedFile or BytesIO対応）
    if isinstance(schedule_file, (str, os.PathLike)):
        with open(schedule_file, "r", encoding="utf-8") as f:
            lines = [line.rstrip("\n").split(",") for line in f]
    else:
        text = schedule_file.read().decode("utf-8")
        lines = [line.rstrip("\n").split(",") for line in text.splitlines()]

    max_cols = max(len(line) for line in lines)
    lines = [line + [""] * (max_cols - len(line)) for line in lines]
    sched = pd.DataFrame(lines).fillna("")

    # PREF ファイル読み込み（str or BytesIO 対応）
    pref = (
        pref_file
        if isinstance(pref_file, (str, os.PathLike))
        else io.BytesIO(pref_file.read())
    )
    emp_df = pd.read_excel(pref, sheet_name="emp_no", header=None, dtype=str).fillna("")

    # 色分けルールと SIM Slot コード読み込み（補助関数が外にある想定）
    pref_rules = load_pref_rules(pref)
    simslot_codes = load_simslot_schedule_codes(pref)

    # SIM Slot ファイル読み込み（str or BytesIO 対応）
    sim = (
        sim_file
        if isinstance(sim_file, (str, os.PathLike))
        else io.BytesIO(sim_file.read())
    )
    wb_sim = load_workbook(sim, read_only=True, data_only=True)
    first_sheet = wb_sim.sheetnames[0]
    sim_df = pd.read_excel(sim, sheet_name=first_sheet, header=2, dtype=str).fillna("")

    # ── 2) 社員情報マップ作成 ─────────────────────────────────────────
    # emp_df から各種マップを生成: 氏名、2レター、所属、背景設定列
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()
    # 養成期 (D 列) と Phase (H 列) のマッピング
    phase_d_map = {
        str(row[2]).strip(): str(row[3]).strip() for _, row in emp_df.iterrows()
    }
    phase_h_map = {
        str(row[2]).strip(): str(row[7]).strip() for _, row in emp_df.iterrows()
    }

    # ── 3) 元データ整形 ────────────────────────────────────────────
    # セルクリーンと空行・OB 行削除
    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)

    # OBなどヘッダーだけでスケジュールが存在しない行を除外（NEW）
    def remove_ob_only_headers(df):
        def is_valid_header(idx):
            if idx + 1 >= len(df):  # 次行が存在しない
                return False
            next_row = df.iloc[idx + 1]
            # 次の行に日付らしき01〜31が5つ以上あるか
            has_date = (
                next_row.astype(str).str.match(r"^(0[1-9]|[12][0-9]|3[01])$").sum() >= 5
            )
            return has_date

        # ヘッダーだけの行や列数が極端に少ない行を除外
        mask = [is_valid_header(i) or df.iloc[i].count() > 30 for i in range(len(df))]
        return df[mask]

    # ↑ 上記のフィルタ関数を適用
    df = remove_ob_only_headers(df)

    # ブロック抽出
    blocks = slice_blocks(df)
    if not blocks:
        return
    global_dates = blocks[0][3]

    # ── 4) SIM Slot 実績読み込み & 参加者辞書作成 ────────────────────────
    from datetime import datetime

    # Excelブックを開いて1つ目のシート名を取得
    wb_sim = load_workbook(sim_file, read_only=True, data_only=True)
    first_sheet = wb_sim.sheetnames[0]

    # 3行目（index=2）を列名、4行目以降をデータとして読み込む
    sim_df = pd.read_excel(
        sim_file, sheet_name=first_sheet, header=2, dtype=str
    ).fillna("")

    # logger.info(f"SIM Slot 列名一覧: {list(sim_df.columns)}")
    # logger.info(f"SIM Slot 読み込み行数: {len(sim_df)}")
    # DEBUG: logger.debug(f"SIM Slot 先頭5行:\n{sim_df.head()}")

    # 列名の標準化： '日付' → 'day'
    if "日付" in sim_df.columns:
        sim_df.rename(columns={"日付": "day"}, inplace=True)
        # print("✔ '日付' 列を 'day' にリネームしました")
    else:
        print("❌ '日付' 列が見つかりません")
        sim_df["day"] = pd.NA  # フォールバック定義

    # Event Name列（存在しない場合もある）
    sim_df["Event Name"] = sim_df.get("Event Name", "").astype(str).str.strip()

    # 'day'列をdatetimeに変換し日部分だけを抽出（Int64でNAも保持）
    if "day" in sim_df.columns:
        try:
            sim_df["day"] = pd.to_datetime(
                sim_df["day"], errors="coerce"
            ).dt.day.astype("Int64")
        except Exception:
            pass  # 既にInt64変換済みならスルー

    # ID列をリストに変換（"/" 区切り）
    sim_df["教官 Emp ID"] = sim_df.get("教官 Emp ID", "").str.split("/", expand=False)
    sim_df["訓練生 Emp ID"] = sim_df.get("訓練生 Emp ID", "").str.split(
        "/", expand=False
    )

    # A350: 空欄 → APT ／ 787: '1','2','7','8' → #番号 に変換する補助関数
    def get_machine(raw):
        if pd.isna(raw) or str(raw).strip() == "":
            return "APT"
        raw = str(raw).strip()
        if raw in {"1", "2", "7", "8"}:
            return f"#{raw}"
        return raw

    # ActivityTypeCode列の存在チェックとフィルタリング
    if "ActivityTypeCode" not in sim_df.columns:
        raise ValueError("SIMファイルに 'ActivityTypeCode' 列が存在しません。")

    # logger.info(f"フィルタ前 SIM Slot 件数: {len(sim_df)}")
    # logger.info(f"フィルタ条件: {simslot_codes}")

    # フィルタ適用
    sim_df = sim_df[sim_df["ActivityTypeCode"].isin(simslot_codes)]

    # logger.info(f"フィルタ後 SIM Slot 件数: {len(sim_df)}")

    # 号機の列は "号機" or "機番" など変動する場合はB列（index=1）を使う
    machine_col = sim_df.columns[1]

    # SIM参加者辞書と各参加者→時間帯辞書を作成
    teacher_lookup = {}
    trainee_lookup = {}
    simslot_participants = {}

    for _, row in sim_df.iterrows():
        day = row.get("day")
        act_code = row.get("ActivityTypeCode", "").strip()
        evt_code = row.get("Event Name", "").strip()
        start = row.get("開始時刻", "")
        end = row.get("終了時刻", "")
        machine = get_machine(row.get(machine_col, ""))

        # ID整形（"/"区切り後の個別IDを5桁へ）
        teachers = [
            eid.strip()[-5:] for eid in row.get("教官 Emp ID", []) if eid.strip()
        ]
        trainees = [
            eid.strip()[-5:] for eid in row.get("訓練生 Emp ID", []) if eid.strip()
        ]

        # 各IDと時間帯を lookup に登録（同一day + emp_id に対して上書き）
        for eid in teachers:
            teacher_lookup[(day, eid)] = (start, end)
        for eid in trainees:
            trainee_lookup[(day, eid)] = (start, end)

        # 参加者辞書に登録（act_code単独と、act+eventが異なる場合はeventもキーに）
        codes = [act_code] + ([evt_code] if evt_code and evt_code != act_code else [])
        for code_key in codes:
            key = (day, code_key, start, end, machine)
            simslot_participants.setdefault(
                key, {"teachers": [], "trainees": [], "event": evt_code}
            )
            simslot_participants[key]["teachers"].extend(teachers)
            simslot_participants[key]["trainees"].extend(trainees)

    # logger.info(f"教官辞書数: {len(teacher_lookup)}")
    # logger.info(f"訓練生辞書数: {len(trainee_lookup)}")
    # logger.info(f"参加者辞書数: {len(simslot_participants)}")

    # 代表例出力（最初の3件）
    for i, (key, val) in enumerate(simslot_participants.items()):
        if i >= 3:
            break
        # logger.debug(
        #     f"[sample {i}] KEY={key}, TEACHERS={val['teachers']}, TRAINEES={val['trainees']}, EVENT={val['event']}"
        # )

    # 重複排除
    for grp in simslot_participants.values():
        grp["teachers"] = list(dict.fromkeys(grp["teachers"]))

    # ── 5) records 作成 ─────────────────────────────────────────────────
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000\d{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        # 元のヘッダー（MATSUNAMI など）を保持
        orig_name = raw[0]

        raw[0] = f"{surname}{two}" if matched else raw[0]

        hdr_vals = [v for v in raw if v]
        hdr = hdr_vals[:31] + [""] * (31 - len(hdr_vals[:31]))
        # 養成期/Phase 表示を設定
        raw_d = phase_d_map.get(code, "").strip()
        raw_h = phase_h_map.get(code, "").strip()
        if raw_d and not raw_d.isdigit():
            # D列がアルファベット系のとき
            if raw_d == "B":
                display = "B教官"
            elif raw_d == "BS":
                display = "S教官"
            elif raw_d == "BSL":
                display = "L教官"
            elif raw_d == "VBSL":
                display = "VLD"
            else:
                display = raw_d
        else:
            # D列が数字のみ or 空文字 → H列を参照
            if raw_h.isdigit():
                display = f"PH{raw_h}"
            else:
                display = raw_h

        hdr[29] = display
        hdr[30] = rec_aff

        # ── セル内文字列整形ルール (まとめて置換) ───────────────────────────────────
        # 「PE有効期限」→「PE」, 「PExxxxxx」→「xxxxxx」
        # 「社員番号」→「職番」, 「電話番号」→「電話」
        # 「CAT資格」→「CAT資」, 「T/O期限」→「T/O」, 「L/D期限」→「L/D」
        hdr = [
            re.sub(
                r"L/D期限",
                "L/D",
                re.sub(
                    r"T/O期限",
                    "T/O",
                    re.sub(
                        r"CAT資格",
                        "CAT資",
                        re.sub(
                            r"電話番号",
                            "電話",
                            re.sub(
                                r"社員番号",
                                "職番",
                                re.sub(
                                    r"PE(\\d{6})",
                                    r"\\1",
                                    re.sub(r"PE有効期限", "PE", v),
                                ),
                            ),
                        ),
                    ),
                ),
            )
            for v in hdr
        ]

        # 日付行・スケジュール行の取得（既存ロジック）
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            fe.append(
                [
                    clean_cell(df.iat[r2, j])
                    for r2 in range(d + 1, end)
                    if clean_cell(df.iat[r2, j])
                ]
            )
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))

        # レコードに追加
        records.append(
            {
                "emp_no": code,
                "orig_name": orig_name,  # ★ ここを追加 ★
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )

    # 全日数埋め
    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))

    # ── 6) Phase1: 訓練コード＋ISR/TRN のみ残す ──────────────────────────────
    for rec in records:
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            first = cell_text.split("\n", 1)[0].strip()
            if first in simslot_codes:
                if "ISR" in cell_text:
                    rec["sched"][idx] = f"{first}\nISR"
                elif "TRN" in cell_text:
                    rec["sched"][idx] = f"{first}\nTRN"

    # ── 7) Phase2: 時間帯＋号機＋Event Name（特定コードのみ）追記 ─────────
    for rec in records:
        emp_id = rec["emp_no"]
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            lines = cell_text.split("\n")
            code = lines[0].strip()
            if code not in simslot_codes or len(lines) < 2:
                continue
            subcode = lines[1].strip()
            if subcode not in ("ISR", "TRN"):
                continue

            # 日付取得
            try:
                day = int(rec["dr"][idx])
            except ValueError:
                continue

            # 時刻取得
            lookup = teacher_lookup if subcode == "ISR" else trainee_lookup
            time_pair = lookup.get((day, emp_id))
            if not time_pair:
                continue
            start, end = time_pair

            # 号機特定
            machine = None
            parts = None
            for (d, c, s, e, m), grp in simslot_participants.items():
                if (d, c, s, e) == (day, code, start, end) and emp_id in grp[
                    "teachers"
                ] + grp["trainees"]:
                    machine = m
                    parts = grp
                    break
            if not parts:
                continue

            # Event Name は特定コードのみ
            event_name = parts.get("event", "")
            base = f"{code}\n{subcode}\n{machine}\n{start}-{end}"
            if code in ("FOTR", "CATR", "CAUG") and event_name:
                rec["sched"][idx] = f"{base}\n{event_name}"
            else:
                rec["sched"][idx] = base
    # ───────────────────────────────────────────────────────────────────

    # ── 8) Phase3: onb ロジック統合（自己除外強化版） ────────────────────────
    for i, rec in enumerate(records):
        emp_id = rec["emp_no"]
        onb = []
        for idx, sched_cell in enumerate(rec["sched"]):
            if not sched_cell:
                onb.append([])
                continue
            lines = sched_cell.split("\n", 1)
            first = lines[0].strip()

            if first in simslot_codes:
                try:
                    day = int(rec["dr"][idx])
                except ValueError:
                    onb.append([])
                    continue

                parts = None
                machine = None

                # Debugログ：マッチング試行を記録
                # logging.debug(
                #     f"🔍 SIM マッチチェック: day={day}, code={first}, emp_id={emp_id}"
                # )

                for (d, c, s, e, m), grp in simslot_participants.items():
                    if (
                        d == day
                        and c == first
                        and emp_id in grp["teachers"] + grp["trainees"]
                    ):
                        parts = grp
                        machine = m
                        # logging.debug(
                        #     f"✅ MATCH: key=({d}, {c}, {s}, {e}, {m}), emp_id={emp_id}"
                        # )
                        break
                    # else:
                    #     logging.debug(
                    #         f"… NO match: key=({d}, {c}, {s}, {e}, {m}), emp_id={emp_id} not in {grp['teachers'] + grp['trainees']}"
                    #     )

                if not parts:
                    # logging.debug(
                    #     f"❌ NO MATCH FOUND for emp_id={emp_id}, code={first}, day={day}"
                    # )
                    onb.append([])
                    continue

                names = []
                for tid in parts["teachers"]:
                    if tid.strip() == emp_id.strip():
                        continue
                    nm = emp_name_map.get(tid, "")
                    two = emp_two_map.get(tid, "")
                    if nm:
                        names.append(f"I{nm}{two}")
                for tid in parts["trainees"]:
                    if tid.strip() == emp_id.strip():
                        continue
                    nm = emp_name_map.get(tid, "")
                    two = emp_two_map.get(tid, "")
                    if nm:
                        names.append(f"T{nm}{two}")
                onb.append(names)
            else:
                entries = rec["full_entries"][idx]
                flights = [e for e in entries if e and re.match(r"^[0-9]", e)]
                names = []
                for j, other in enumerate(records):
                    if i == j or other["hdr"][0] == rec["hdr"][0]:
                        continue
                    other_entries = other["full_entries"][idx]
                    if any(f in other_entries for f in flights):
                        names.append(other["hdr"][0])
                uniq = []
                for n in names:
                    if n not in uniq:
                        uniq.append(n)
                onb.append(uniq)
        rec["onb"] = onb

    # ── 9) 重複レコード削除＆ソート ────────────────────────────────────────────
    seen, uniq = set(), []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = sorted(
        uniq,
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        ),
    )

    # 日付文字列取得 (Asia/Tokyo)
    today = datetime.now().strftime("%Y%m%d")
    out_csv = f"SKDFILE{today}.csv"
    out_xlsx = f"SKDFILE{today}.xlsx"

    # CSV 出力
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec.get("onb", [])])

    # Excel 出力
    write_to_excel(records, emp_aff_map, out_xlsx, pref_rules)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--pref", default="PREF.xlsx")
    p.add_argument(
        "--sim", default="SIM Slot List 202507.xlsx"
    )  # SIM Slot ファイルをオプションとして追加

    a = p.parse_args()
    run(a.schedule, a.pref, a.sim)  # ← sim を明示的に渡す
