# === generate_schedule.py ===
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# ==== Helpers ====


def clean_cell(text):
    """Remove invisible characters and trim whitespace"""
    s = str(text)
    return re.sub(r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", s).strip()


def remove_blank_and_ob(df):
    """Drop rows that are all blank or contain only 'OB'"""
    df = df.replace({pd.NA: "", "nan": "", "NaN": ""})
    df = df[(df != "").any(axis=1)]
    df = df[~df.apply(lambda row: all(str(c).strip() == "OB" for c in row), axis=1)]
    return df.reset_index(drop=True)


def find_header_rows(df):
    """
    Detect header rows where:
    - first cell ends with two uppercase letters
    - the next row contains exactly 31 date values (1-31)
    """
    hdrs = []
    for i in range(len(df) - 1):
        if re.search(r"[A-Z]{2}$", clean_cell(df.iat[i, 0])):
            next_row = [clean_cell(x) for x in df.iloc[i + 1]]
            dates = [
                v for v in next_row if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
            ]
            if len(dates) == 31:
                hdrs.append(i)
    return hdrs


def slice_blocks(df):
    hdrs = find_header_rows(df)
    blocks = []
    total_rows = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total_rows
        # Find date row
        for r in range(h + 1, end):
            row = [clean_cell(x) for x in df.iloc[r]]
            date_positions = [
                j
                for j, v in enumerate(row)
                if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
            ]
            if date_positions:
                # only first date row
                blocks.append((h, r, end, date_positions))
                break
    return blocks


def write_to_excel(records, out_xlsx):
    wb = Workbook()
    ws = wb.active
    grey = PatternFill(fill_type="solid", fgColor="DDDDDD")
    hi = PatternFill(fill_type="solid", fgColor="FFFF00")
    dbl = Side(border_style="double", color="000000")
    brd = Border(top=dbl, bottom=dbl)

    # Set column widths
    for c in range(1, 32):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 8

    row_num = 1
    for rec in records:
        # Header row
        for c, v in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.border = brd
        row_num += 1

        # Date row
        for c, v in enumerate(rec["dr"], start=1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = grey
        row_num += 1

        # Schedule row
        for c, v in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        row_num += 1

        # Onboard row
        for c, v in enumerate(rec["onb"], start=1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if "*" in v:
                cell.fill = hi
        row_num += 1

    wb.save(out_xlsx)


def run(schedule_input, emp_input, config_path=None):
    # Load data
    sched_df = pd.read_csv(schedule_input, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_input, header=None, dtype=str).fillna("")

    # Build employee map: code (5 digits) -> combined name
    emp_map = {}
    for _, row in emp_df.iterrows():
        code = str(row[2]).zfill(5)
        if re.fullmatch(r"[0-9]{5}", code):
            part1 = str(row[4]) if pd.notna(row[4]) else ""
            part2 = str(row[6]) if pd.notna(row[6]) else ""
            emp_map[code] = part1 + part2

    # Clean and block slice
    df = sched_df.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)

    records = []
    for h, d, end, dates in blocks:
        # Header: remove blanks, left-align, pad to 31
        raw = [clean_cell(x) for x in df.iloc[h]]
        hdr_vals = [v for v in raw if v]
        hdr = hdr_vals[:31] + [""] * (31 - len(hdr_vals[:31]))

        # Replace name if code matches
        if len(raw) > 8:
            m = re.search(r"000([0-9]{5})", raw[8])
            if m:
                code = m.group(1)
                if code in emp_map:
                    hdr[0] = emp_map[code]

        # Date row values
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))

        # Schedule entries
        full_entries = [
            [
                clean_cell(df.iat[r, j])
                for r in range(d + 1, end)
                if clean_cell(df.iat[r, j])
            ]
            for j in dates
        ]
        sched = ["\n".join(entries) for entries in full_entries] + [""] * (
            31 - len(full_entries)
        )

        # Onboard names
        onb = []
        for idx, flights in enumerate(full_entries):
            names = []
            for k, other in enumerate(records):
                other_flights = (
                    other["sched"][idx].split("\n") if idx < len(other["sched"]) else []
                )
                if any(f in other_flights for f in flights if re.match(r"^[0-9]", f)):
                    names.append(other["hdr"][0])
            onb.append("\n".join(names))
        onb += [""] * (31 - len(onb))

        records.append({"hdr": hdr, "dr": dr, "sched": sched, "onb": onb})

    # CSV output
    out_csv = "formatted_schedule.csv"
    rows = []
    for rec in records:
        rows.extend([rec["hdr"], rec["dr"]] + rec["sched"] + [rec["onb"]])
    pd.DataFrame(rows).to_csv(out_csv, index=False, header=False)

    # Excel output
    out_xlsx = "formatted_schedule.xlsx"
    write_to_excel(records, out_xlsx)

    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--schedule", default="schedule.csv")
    parser.add_argument("--emp", default="emp_no.csv")
    args = parser.parse_args()
    run(args.schedule, args.emp)
