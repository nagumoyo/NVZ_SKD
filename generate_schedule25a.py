#!/usr/bin/env python3
# === generate_schedule24.py ===

import pandas as pd
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==== Helpers ====


def clean_cell(text):
    return re.sub(
        r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", str(text)
    ).strip()


def remove_blank_and_ob(df):
    rows = []
    for i, row in enumerate(df.values):
        texts = [str(x).strip() for x in row]
        if any(re.fullmatch(r"00099[0-9]{3}", t) for t in texts):
            continue
        if all(not t or t == "OB" for t in texts):
            continue
        if re.fullmatch(r"[A-Z]+OB", texts[0]):
            continue
        rows.append(row)
    return pd.DataFrame(rows, columns=df.columns)


def find_header_rows(df):
    hdrs = []
    for i in range(len(df) - 1):
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    return hdrs


def slice_blocks(df):
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        d = h + 1
        vals = [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    d = r
                    break
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]
        blocks.append((h, d, end, date_cols))
    return blocks


def load_pref_rules(pref_file):
    from openpyxl import load_workbook
    import io

    # ファイルパスかファイルオブジェクトかを判定
    if isinstance(pref_file, str):
        wb = load_workbook(filename=pref_file)
    else:
        pref_file.seek(0)  # 読み込み位置を先頭に戻す
        wb = load_workbook(filename=io.BytesIO(pref_file.read()))

    ws = wb.active
    data = list(ws.iter_rows(min_row=2, values_only=True))
    rules = []
    for i, row in enumerate(data):
        enable, first, op, second, label, _ = row[:6]
        cell = ws.cell(row=i + 2, column=6)
        fill = cell.fill
        color = (
            fill.start_color.rgb
            if fill and fill.start_color and fill.start_color.type == "rgb"
            else ""
        )
        rgb_color = f"#{color[-6:]}" if color else ""
        if str(enable).strip().upper() == "YES":
            rules.append(
                {
                    "first": str(first).strip(),
                    "second": str(second).strip(),
                    "op": str(op).strip().upper(),
                    "label": str(label).strip(),
                    "color": rgb_color,
                }
            )
    return rules


def apply_pref_rules_to_cell(cell, val, rules, fallback_color=None):
    import re

    text = str(val)
    applied = False
    for rule in rules:
        try:
            cond1 = bool(re.search(rule["first"], text)) if rule["first"] else False
            cond2 = bool(re.search(rule["second"], text)) if rule["second"] else False
        except re.error:
            continue
        if rule["op"] == "AND" and cond1 and cond2:
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            applied = True
            break
        elif rule["op"] == "OR" and (cond1 or cond2):
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            applied = True
            break
        elif rule["op"] not in ("AND", "OR") and cond1:
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            applied = True
            break
    if not applied and fallback_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fallback_color)


def write_onboard_rows(
    ws,
    start_row,
    onboard_data,
    emp_aff_map,
    name_to_emp,
    block_aff,
    self_name,
    name_to_row,
):
    max_onb = max((len(day) for day in onboard_data if day), default=1)
    for i in range(max_onb):
        for j, names in enumerate(onboard_data, start=1):
            value = names[i] if i < len(names) else ""
            if value == self_name:
                value = ""
            target_row = name_to_row.get(value)
            cell = ws.cell(row=start_row + i, column=j)
            if value and target_row:
                cell.value = f'=HYPERLINK("#A{target_row}", "{value}")'
            else:
                cell.value = value
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if value:
                emp = name_to_emp.get(value)
                if emp and emp_aff_map.get(emp) == block_aff:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFEE99")
    return max_onb


def write_to_excel(records, emp_aff_map, out_xlsx, pref_rules):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}
    name_to_row = {}
    row_counter = 1
    for rec in records:
        row_counter += 3 + max((len(x) for x in rec.get("onb", [])), default=1)
        name_to_row[rec["hdr"][0]] = row_counter

    row_num = 1
    for rec in records:
        block_aff = rec["aff"]
        self_name = rec["hdr"][0]

        for j, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=j, value=val)
            wrap = not bool(re.fullmatch(r"0[0-9]{1,}-[0-9]+-[0-9]{4}", val))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=wrap
            )
            cell.border = Border(
                top=Side(border_style="double", color="000000"),
                bottom=Side(border_style="double", color="000000"),
            )

        for j, date_val in enumerate(rec["dr"], start=1):
            sched_val = rec["sched"][j - 1] if j - 1 < len(rec["sched"]) else ""
            cell = ws.cell(row=row_num + 1, column=j, value=date_val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            apply_pref_rules_to_cell(
                cell, sched_val, pref_rules, fallback_color="DDDDDD"
            )

        for j, val in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num + 2, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

        onboard_count = write_onboard_rows(
            ws,
            row_num + 3,
            rec.get("onb", []),
            emp_aff_map,
            name_to_emp,
            block_aff,
            self_name,
            name_to_row,
        )
        row_num += 3 + onboard_count

    wb.save(out_xlsx)


# その他 main 関数などは既存通り（適宜 pref_rules を渡すようにする）


def run(schedule_file, emp_file, pref_file="PREF.xlsx"):
    sched = pd.read_csv(schedule_file, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_file, header=None, dtype=str).fillna("")
    pref_rules = load_pref_rules(pref_file)
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()
    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)
    if not blocks:
        return
    global_dates = blocks[0][3]
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        raw[0] = f"{surname}{two}" if matched else raw[0]
        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        col8 = emp_col8_map.get(code, "")
        m = re.search(r"(\d+)", col8)
        if m:
            hdr[29] = f"PH{m.group(1)}"
        hdr[30] = rec_aff
        hdr = [
            re.sub(
                r"電話番号",
                "電話",
                re.sub(
                    r"社員番号",
                    "職番",
                    re.sub(r"PE([0-9]{6})", r"\1", re.sub(r"PE有効期限", "PE", v)),
                ),
            )
            for v in hdr
        ]
        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            fe.append(
                [
                    clean_cell(df.iat[r2, j])
                    for r2 in range(d + 1, end)
                    if clean_cell(df.iat[r2, j])
                ]
            )
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))
        records.append(
            {
                "emp_no": code,
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )
    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))
    for i, rec in enumerate(records):
        onb = []
        for idx, entries in enumerate(rec["full_entries"]):
            flights = [e for e in entries if e and re.match(r"^[0-9]", e)]
            names = []
            for j, other in enumerate(records):
                if i == j:
                    continue
                if any(f in other["full_entries"][idx] for f in flights):
                    names.append(other["hdr"][0])
            uniq = []
            for n in names:
                if n not in uniq:
                    uniq.append(n)
            onb.append(uniq)
        rec["onb"] = onb
    seen = set()
    uniq = []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = uniq
    records.sort(
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        )
    )
    out_csv = "formatted_schedule.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec["onb"]])
    out_xlsx = "formatted_schedule20.xlsx"
    write_to_excel(records, emp_aff_map, out_xlsx, pref_rules)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--emp", default="emp_no.csv")
    p.add_argument("--pref", default="PREF.xlsx")
    a = p.parse_args()
    run(a.schedule, a.emp, a.pref)
