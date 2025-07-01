#!/usr/bin/env python3
# === generate_schedule24.py ===

import pandas as pd
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import warnings

warnings.filterwarnings(
    "ignore", category=UserWarning, module="openpyxl.styles.stylesheet"
)

# --- SIMSLOTシートから対象社員IDリストを取得 ---
wb_pref = load_workbook("PREF.xlsx", data_only=True)
ws_sim = wb_pref["SIMSLOT"]

# B1 セルの値を取得（例えば "00012345,00067890,00054321" のようなカンマ区切り想定）
b1_value = ws_sim["B1"].value

# 文字列をカンマで分割し、空白をトリムしたリストに変換
target_codes = [emp.strip() for emp in str(b1_value).split(",") if emp.strip()]

# デバッグ用：取得結果を確認
# print(f"SIMSLOT 対象社員 ID リスト: {target_emp_ids}")

import pandas as pd

# ─── ② SIM Slot 実績データ読み込み & 辞書化 ───────────────────────────────
sim_df = pd.read_excel(
    "SIM Slot List 202507.xlsx", sheet_name="SIM Slot List", header=2, dtype=str
).fillna("")

# 「ActivityTypeCode」が target_codes（B1リスト）に含まれない行は除外
sim_df = sim_df[sim_df["ActivityTypeCode"].isin(target_codes)]

# 日付→day列作成
sim_df["日付"] = pd.to_datetime(sim_df["日付"], format="%Y/%m/%d", errors="coerce")
sim_df["day"] = sim_df["日付"].dt.day.astype("Int64")

# Emp ID列をリスト化
sim_df["教官 Emp ID"] = sim_df["教官 Emp ID"].str.split("/", expand=False)
sim_df["訓練生 Emp ID"] = sim_df["訓練生 Emp ID"].str.split("/", expand=False)

# 辞書型ルックアップテーブルを生成（空文字はスキップ）
teacher_lookup = {}
trainee_lookup = {}

for _, row in sim_df.iterrows():
    d = int(row["day"])
    start = row["開始時刻"]
    end = row["終了時刻"]

    for eid in row["教官 Emp ID"]:
        eid = eid.strip()
        if not eid:
            continue
        # フル8桁から末尾5桁だけ取り出してキーにする
        emp_key = eid[-5:]
        teacher_lookup[(d, emp_key)] = (start, end)

    for eid in row["訓練生 Emp ID"]:
        eid = eid.strip()
        if not eid:
            continue
        emp_key = eid[-5:]
        trainee_lookup[(d, emp_key)] = (start, end)
# ── (SIM Slot 読み込み部の末尾に追記) ─────────────────────────────
# sim_df: ActivityTypeCode, day, 教官 Emp ID リスト, 訓練生 Emp ID リスト がある DataFrame

# 参加者リスト辞書を作成
# key: (day, training_code), value: {"teachers": [...], "trainees": [...]}
simslot_participants = {}
for _, row in sim_df.iterrows():
    code = row["ActivityTypeCode"]
    day = int(row["day"])
    key = (day, code)
    # 教官側／訓練生側のリストはすでに row["教官 Emp ID"] が ['00012',...] 末尾5桁で分割済み、
    # もしくは str.split() 後に末尾5桁化しているはずなのでそのまま使えます。
    teachers = [eid.strip()[-5:] for eid in row["教官 Emp ID"] if eid.strip()]
    trainees = [eid.strip()[-5:] for eid in row["訓練生 Emp ID"] if eid.strip()]
    simslot_participants[key] = {"teachers": teachers, "trainees": trainees}

# 動作確認用
# print("教師ルックアップ件数:", len(teacher_lookup))
# print("訓練生ルックアップ件数:", len(trainee_lookup))


# ==== Helpers ====


def load_sim_slot_excel(file):
    """
    SIM Slot ExcelファイルをDataFrameとして読み込む（Streamlitまたはローカル兼用）
    :param file: ファイルパスまたはStreamlitのアップロードファイルオブジェクト
    :return: pandas.DataFrame
    """
    import pandas as pd
    import io

    try:
        if isinstance(file, str):
            # ローカルファイルとして読み込み
            df = pd.read_excel(file, header=2)
        else:
            # Streamlitアップロードファイルを読み込み
            df = pd.read_excel(io.BytesIO(file.read()), header=2)

        print(f"[INFO] SIM Slot: {len(df)} 行を読み込みました")
        return df

    except Exception as e:
        print(f"[ERROR] SIM Slot Excel読込失敗: {e}")
        return pd.DataFrame()


def load_simslot_schedule_codes(pref_file_path="PREF.xlsx", sheet_name="SIMSLOT"):
    """
    PREF.xlsx の SIMSLOT シートから B1 にあるカンマ区切りのコード群を取得し、リストとして返す。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(pref_file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"✅ SIMSLOT シートが存在しません: {sheet_name}")
            return []

        ws = wb[sheet_name]
        b1_value = ws["B1"].value
        if not b1_value:
            print("✅ B1セルが空です")
            return []

        return [code.strip() for code in str(b1_value).split(",") if code.strip()]

    except Exception as e:
        print(f"⚠️ SIMSLOT 読み込み失敗: {e}")
        return []


def clean_cell(text):
    return re.sub(
        r"[\u200b\u200c\u200d\u2060\uFEFF\u00A0\t\r\n]", "", str(text)
    ).strip()


def remove_blank_and_ob(df):
    rows = []
    for i, row in enumerate(df.values):
        texts = [str(x).strip() for x in row]
        if any(re.fullmatch(r"00099[0-9]{3}", t) for t in texts):
            continue
        if all(not t or t == "OB" for t in texts):
            continue
        if re.fullmatch(r"[A-Z]+OB", texts[0]):
            continue
        rows.append(row)
    return pd.DataFrame(rows, columns=df.columns)


def find_header_rows(df):
    hdrs = []
    for i in range(len(df) - 1):
        c0 = str(df.iat[i, 0]).strip()
        c2 = str(df.iat[i, 2]).strip() if df.shape[1] > 2 else ""
        next_row = [str(df.iat[i + 1, j]).strip() for j in range(df.shape[1])]
        if (
            re.fullmatch(r"[A-Z]{2,}", c0)
            and re.fullmatch(r"[A-Z]{2}", c2)
            and any(
                re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in next_row if v
            )
        ):
            hdrs.append(i)
    return hdrs


def slice_blocks(df):
    hdrs = find_header_rows(df)
    blocks = []
    total = len(df)
    for idx, h in enumerate(hdrs):
        end = hdrs[idx + 1] if idx + 1 < len(hdrs) else total
        d = h + 1
        vals = [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
        if not any(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in vals if v):
            for r in range(h + 1, end):
                tmp = [str(df.iat[r, j]).strip() for j in range(df.shape[1])]
                if all(re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v) for v in tmp if v):
                    d = r
                    break
        date_cols = [
            j
            for j, v in enumerate(
                [str(df.iat[d, j]).strip() for j in range(df.shape[1])]
            )
            if re.fullmatch(r"(0?[1-9]|[12][0-9]|3[01])", v)
        ]
        blocks.append((h, d, end, date_cols))
    return blocks


def load_pref_rules(file="PREF.xlsx", sheet_name="色分け設定"):
    from openpyxl import load_workbook

    wb = load_workbook(file)

    # 指定されたシートが存在しない場合は空リストを返す
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] シート「{sheet_name}」が {file} に存在しません。")
        return []

    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=2, values_only=True))
    rules = []

    for i, row in enumerate(data):
        enable, first, op, second, label, _ = row[:6]
        cell = ws.cell(row=i + 2, column=6)  # 6列目の色取得
        fill = cell.fill
        color = (
            fill.start_color.rgb
            if fill and fill.start_color and fill.start_color.type == "rgb"
            else ""
        )
        rgb_color = f"#{color[-6:]}" if color else ""
        if str(enable).strip().upper() == "YES":
            rules.append(
                {
                    "first": str(first or "").strip(),
                    "second": str(second or "").strip(),
                    "op": str(op or "NONE").strip().upper(),
                    "label": str(label or "").strip(),
                    "color": rgb_color,
                }
            )
    return rules


def match_rule_in_multiline(text, rule, debug_log_path="debug_log.txt"):
    lines = str(text).split("\n")
    cond1 = rule["first"].strip()
    cond2 = rule["second"].strip()
    op = rule["op"].strip().upper()

    cond1_found = any(cond1 in line for line in lines) if cond1 else False
    cond2_found = any(cond2 in line for line in lines) if cond2 else False

    # ログをファイルに追記
    with open(debug_log_path, "a", encoding="utf-8") as f:
        f.write("\n=== DEBUG MATCH ===\n")
        f.write(f"Text:\n{text}\n")
        f.write(f"Lines: {lines}\n")
        f.write(f"Rule: {rule}\n")
        f.write(f"cond1_found: {cond1_found}, cond2_found: {cond2_found}, op: {op}\n")

    if op == "AND":
        return cond1_found and cond2_found
    elif op == "OR":
        return cond1_found or cond2_found
    else:
        return False


def apply_pref_rules_to_cell(cell, val, rules, fallback_color=None):
    """
    PREFルールに基づいてセルの背景色を設定。
    スケジュール文字列が複数行ある場合は行ごとに正規表現判定する。
    """
    text = str(val).strip()
    lines = [line.strip() for line in text.split("\n") if line.strip()]

    for idx, rule in enumerate(rules):
        if not isinstance(rule, dict):
            print(f"⚠️ Rule #{idx} is not a dict: {rule} (type={type(rule)})")
            continue

        first = rule.get("first", "") or ""
        second = rule.get("second", "") or ""
        op = rule.get("op", "NONE").upper()

        # 行単位で正規表現マッチ確認
        cond1_found = (
            any(re.search(first, line) for line in lines)
            if first and first != "None"
            else False
        )
        cond2_found = (
            any(re.search(second, line) for line in lines)
            if second and second != "None"
            else False
        )

        if op == "AND" and cond1_found and cond2_found:
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            return
        elif op == "OR" and (cond1_found or cond2_found):
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            return
        elif op == "NONE" and (cond1_found or cond2_found):
            cell.fill = PatternFill(
                fill_type="solid", fgColor=rule["color"].replace("#", "")
            )
            return

    # どのルールにもマッチしなかった場合のデフォルト色
    if fallback_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fallback_color)


def write_onboard_rows(
    ws,
    start_row,
    onboard_data,
    emp_aff_map,
    name_to_emp,
    block_aff,
    self_name,
    name_to_row,
):
    max_onb = max((len(day) for day in onboard_data if day), default=1)
    for i in range(max_onb):
        for j, names in enumerate(onboard_data, start=1):
            # 元の値（例: "I末継HI" / "T田中AB" / "佐藤CD"）
            display = names[i] if i < len(names) else ""
            cell = ws.cell(row=start_row + i, column=j)

            # プレフィックスを外したキーを探す
            if display.startswith(("I", "T")):
                raw_name = display[1:]  # 先頭1文字(I/T)を削除
            else:
                raw_name = display

            # リンク先行番号を name_to_row から探す
            target_row = name_to_row.get(raw_name)
            if display and target_row:
                # 表示はプレフィックス付き、リンク先は raw_name のスケジュール行
                cell.value = f'=HYPERLINK("#A{target_row}", "{display}")'
            else:
                cell.value = display

            # 既存の書式設定
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            if display:
                emp = name_to_emp.get(raw_name)
                if emp and emp_aff_map.get(emp) == block_aff:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFEE99")
    return max_onb


def write_to_excel(records, emp_aff_map, out_xlsx, pref_rules):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    name_to_emp = {rec["hdr"][0]: rec["emp_no"] for rec in records}
    name_to_row = {}
    row_counter = 1
    for rec in records:
        row_counter += 3 + max((len(x) for x in rec.get("onb", [])), default=1)
        name_to_row[rec["hdr"][0]] = row_counter

    row_num = 1
    for rec in records:
        block_aff = rec["aff"]
        self_name = rec["hdr"][0]

        for j, val in enumerate(rec["hdr"], start=1):
            cell = ws.cell(row=row_num, column=j, value=val)
            wrap = not bool(re.fullmatch(r"0[0-9]{1,}-[0-9]+-[0-9]{4}", val))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=wrap
            )
            cell.border = Border(
                top=Side(border_style="double", color="000000"),
                bottom=Side(border_style="double", color="000000"),
            )

        for j, date_val in enumerate(rec["dr"], start=1):
            sched_val = rec["sched"][j - 1] if j - 1 < len(rec["sched"]) else ""
            cell = ws.cell(row=row_num + 1, column=j, value=date_val)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            apply_pref_rules_to_cell(
                cell, sched_val, pref_rules, fallback_color="DDDDDD"
            )

        for j, val in enumerate(rec["sched"], start=1):
            cell = ws.cell(row=row_num + 2, column=j, value=val)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

        onboard_count = write_onboard_rows(
            ws,
            row_num + 3,
            rec.get("onb", []),
            emp_aff_map,
            name_to_emp,
            block_aff,
            self_name,
            name_to_row,
        )
        row_num += 3 + onboard_count

    wb.save(out_xlsx)


# その他 main 関数などは既存通り（適宜 pref_rules を渡すようにする）


def run(
    schedule_file, emp_file, pref_file="PREF.xlsx", sim_file="SIM Slot List 202507.xlsx"
):
    import pandas as pd
    import re
    import csv
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 1) 入力ファイル読込
    sched = pd.read_csv(schedule_file, header=None, dtype=str).fillna("")
    emp_df = pd.read_csv(emp_file, header=None, dtype=str).fillna("")
    pref_rules = load_pref_rules(pref_file)
    simslot_codes = load_simslot_schedule_codes(pref_file)

    # 2) 社員マップ作成
    emp_name_map = {row[2]: row[4] for _, row in emp_df.iterrows()}
    emp_two_map = {row[2]: row[6] for _, row in emp_df.iterrows()}
    emp_aff_map = {row[2]: row[0] for _, row in emp_df.iterrows()}
    emp_col8_map = {row[2]: row[7] for _, row in emp_df.iterrows()}
    emp_order = emp_df.iloc[:, 2].tolist()

    # 3) 元データ整形
    df = sched.copy().map(clean_cell).pipe(remove_blank_and_ob)
    blocks = slice_blocks(df)
    if not blocks:
        return
    global_dates = blocks[0][3]

    # 4) SIM Slot 実績を読み込み、参加者辞書を作成
    sim_df = pd.read_excel(
        sim_file, sheet_name="SIM Slot List", header=2, dtype=str
    ).fillna("")
    sim_df = sim_df[sim_df["ActivityTypeCode"].isin(simslot_codes)]
    sim_df["日付"] = pd.to_datetime(sim_df["日付"], format="%Y/%m/%d", errors="coerce")
    sim_df["day"] = sim_df["日付"].dt.day.astype(int)
    sim_df["教官 Emp ID"] = sim_df["教官 Emp ID"].str.split("/", expand=False)
    sim_df["訓練生 Emp ID"] = sim_df["訓練生 Emp ID"].str.split("/", expand=False)

    simslot_participants = {}
    for _, row in sim_df.iterrows():
        key = (row["day"], row["ActivityTypeCode"])
        teachers = [eid.strip()[-5:] for eid in row["教官 Emp ID"] if eid.strip()]
        trainees = [eid.strip()[-5:] for eid in row["訓練生 Emp ID"] if eid.strip()]
        simslot_participants[key] = {"teachers": teachers, "trainees": trainees}

    # 5) records 作成
    records = []
    for h, d, end, dates in blocks:
        raw = [clean_cell(x) for x in df.iloc[h]]
        matched = [v for v in raw if re.fullmatch(r"000[0-9]{5}", v)]
        code = matched[0][3:] if matched else ""
        surname = emp_name_map.get(code, clean_cell(df.iat[h, 0]))
        two = emp_two_map.get(code, clean_cell(df.iat[h, 2]))
        rec_aff = emp_aff_map.get(code, "")
        raw[0] = f"{surname}{two}" if matched else raw[0]

        vals = [v for v in raw if v]
        hdr = vals[:31] + [""] * (31 - len(vals[:31]))
        col8 = emp_col8_map.get(code, "")
        m = re.search(r"(\d+.+)", col8)
        if m:
            hdr[29] = f"PH{m.group(1)}"
        hdr[30] = rec_aff
        hdr = [
            re.sub(
                r"電話番号",
                "電話",
                re.sub(
                    r"社員番号",
                    "職番",
                    re.sub(r"PE([0-9]{6})", r"\1", re.sub(r"PE有効期限", "PE", v)),
                ),
            )
            for v in hdr
        ]

        dr = [clean_cell(df.iat[d, j]) for j in dates] + [""] * (31 - len(dates))
        fe = []
        for j in dates:
            fe.append(
                [
                    clean_cell(df.iat[r2, j])
                    for r2 in range(d + 1, end)
                    if clean_cell(df.iat[r2, j])
                ]
            )
        sched_row = ["\n".join(e) for e in fe] + [""] * (31 - len(fe))

        records.append(
            {
                "emp_no": code,
                "hdr": hdr,
                "dr": dr,
                "sched": sched_row,
                "full_entries": fe,
                "aff": rec_aff,
            }
        )

    for rec in records:
        if len(rec["full_entries"]) < len(global_dates):
            rec["full_entries"] += [[]] * (len(global_dates) - len(rec["full_entries"]))

    # 6) Ver26 Phase1: 訓練コード＋ISR/TRN のみ残す
    for rec in records:
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            lines = cell_text.split("\n")
            first = lines[0].strip()
            if first in simslot_codes and "ISR" in cell_text:
                rec["sched"][idx] = f"{first}\nISR"
            elif first in simslot_codes and "TRN" in cell_text:
                rec["sched"][idx] = f"{first}\nTRN"

    # 7) Ver26 Phase2: SIMSLOT 辞書から時間帯を追記
    for rec in records:
        emp_id = rec["emp_no"]
        for idx, cell_text in enumerate(rec["sched"]):
            if not cell_text:
                continue
            if "ISR" in cell_text:
                code = "ISR"
            elif "TRN" in cell_text:
                code = "TRN"
            else:
                continue

            day_str = rec["dr"][idx]
            try:
                day = int(day_str)
            except ValueError:
                continue

            lookup = teacher_lookup if code == "ISR" else trainee_lookup
            time_pair = lookup.get((day, emp_id))
            if not time_pair:
                continue

            start, end = time_pair
            rec["sched"][idx] = f"{cell_text}\n{start}-{end}"

    # 8) Ver27a: onb ロジック統合（訓練／通常フライト両対応）
    for i, rec in enumerate(records):
        emp_id = rec["emp_no"]
        onb = []
        for idx, sched_cell in enumerate(rec["sched"]):
            first_line = sched_cell.split("\n", 1)[0].strip() if sched_cell else ""
            if first_line in simslot_codes:
                # 訓練時の同乗者（I/T）
                day_str = rec["dr"][idx]
                try:
                    day = int(day_str)
                except ValueError:
                    onb.append([])
                    continue

                parts = simslot_participants.get(
                    (day, first_line), {"teachers": [], "trainees": []}
                )
                names = []
                for tid in parts["teachers"]:
                    if tid != emp_id:
                        nm = emp_name_map.get(tid, "")
                        two = emp_two_map.get(tid, "")
                        if nm:
                            names.append(f"I{nm}{two}")
                for tid in parts["trainees"]:
                    if tid != emp_id:
                        nm = emp_name_map.get(tid, "")
                        two = emp_two_map.get(tid, "")
                        if nm:
                            names.append(f"T{nm}{two}")
                onb.append(names)
            else:
                # 通常フライトの同乗者（便番号ベース）
                entries = rec["full_entries"][idx]
                flights = [e for e in entries if e and re.match(r"^[0-9]", e)]
                names = []
                for j, other in enumerate(records):
                    if i == j:
                        continue
                    other_entries = other["full_entries"][idx]
                    if any(f in other_entries for f in flights):
                        names.append(other["hdr"][0])
                uniq = []
                for n in names:
                    if n not in uniq:
                        uniq.append(n)
                onb.append(uniq)
        rec["onb"] = onb

    # 9) 重複レコード削除・ソート
    seen = set()
    uniq = []
    for rec in records:
        key = (rec["emp_no"], tuple(rec["sched"]))
        if key not in seen:
            uniq.append(rec)
            seen.add(key)
    records = uniq
    records.sort(
        key=lambda r: (
            emp_order.index(r["emp_no"]) if r["emp_no"] in emp_order else float("inf")
        )
    )

    # 10) 出力
    out_csv = "formatted_schedule.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for rec in records:
            w.writerow(rec["hdr"])
            w.writerow(rec["dr"])
            w.writerow(rec["sched"])
            w.writerow(["\n".join(x) for x in rec["onb"]])

    out_xlsx = "formatted_schedule20.xlsx"
    write_to_excel(records, emp_aff_map, out_xlsx, pref_rules)
    return out_csv, out_xlsx


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--schedule", default="schedule.csv")
    p.add_argument("--emp", default="emp_no.csv")
    p.add_argument("--pref", default="PREF.xlsx")
    p.add_argument(
        "--sim", default="SIM Slot List 202507.xlsx"
    )  # SIM Slot ファイルをオプションとして追加

    a = p.parse_args()
    run(a.schedule, a.emp, a.pref)
